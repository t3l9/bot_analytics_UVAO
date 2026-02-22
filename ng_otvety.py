from selenium.webdriver import Keys
from telegram import Update, ReplyKeyboardMarkup, InputFile, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes, \
    ConversationHandler, CallbackQueryHandler, CallbackContext
import time
import pandas as pd
import os
from datetime import datetime, timedelta
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
import win32com.client
from functools import reduce
import pythoncom
from dotenv import load_dotenv
from functools import reduce

excluded_dates = [
    # Ваши исходные даты 2025 года
    "20.12.2025", "21.12.2025", "27.12.2025", "28.12.2025", "31.12.2025",

    # Новогодние каникулы и Рождество (2026)
    "01.01.2026", "02.01.2026", "03.01.2026", "04.01.2026", "05.01.2026",
    "06.01.2026", "07.01.2026", "08.01.2026", "09.01.2026", "10.01.2026",
    "11.01.2026",

    # Февральские праздники (23 февраля выпадает на понедельник, добавляем только его)
    "23.02.2026",
    # Предшествующие выходные
    "21.02.2026", "22.02.2026",

    # Мартовские праздники (8 марта - воскресенье, выходной переносится на 9 марта)
    "07.03.2026", "08.03.2026", "09.03.2026",

    # Майские праздники (1-3 мая и 9-11 мая)
    "01.05.2026", "02.05.2026", "03.05.2026",
    "09.05.2026", "10.05.2026", "11.05.2026",

    # День России (12 июня - пятница, длинные выходные 13-14 июня)
    "12.06.2026", "13.06.2026", "14.06.2026",

    # Ноябрьские праздники (4 ноября - среда, отдельный выходной)
    "04.11.2026",
    # Ближайшие выходные
    "31.10.2026", "01.11.2026", "07.11.2026", "08.11.2026",

    # Стандартные выходные 2026 года (субботы и воскресенья, не попавшие в периоды выше)
    # Январь
    "17.01.2026", "18.01.2026", "24.01.2026", "25.01.2026", "31.01.2026",
    # Февраль
    "01.02.2026", "07.02.2026", "08.02.2026", "14.02.2026", "15.02.2026", "28.02.2026",
    # Март
    "01.03.2026", "14.03.2026", "15.03.2026", "21.03.2026", "22.03.2026", "28.03.2026", "29.03.2026",
    # Апрель
    "04.04.2026", "05.04.2026", "11.04.2026", "12.04.2026", "18.04.2026", "19.04.2026", "25.04.2026", "26.04.2026",
    # Май (добавлены только неохваченные)
    "16.05.2026", "17.05.2026", "23.05.2026", "24.05.2026", "30.05.2026", "31.05.2026",
    # Июнь (добавлены только неохваченные)
    "06.06.2026", "07.06.2026", "20.06.2026", "21.06.2026", "27.06.2026", "28.06.2026",
    # Июль
    "04.07.2026", "05.07.2026", "11.07.2026", "12.07.2026", "18.07.2026", "19.07.2026", "25.07.2026", "26.07.2026",
    # Август
    "01.08.2026", "02.08.2026", "08.08.2026", "09.08.2026", "15.08.2026", "16.08.2026", "22.08.2026", "23.08.2026",
    "29.08.2026", "30.08.2026",
    # Сентябрь
    "05.09.2026", "06.09.2026", "12.09.2026", "13.09.2026", "19.09.2026", "20.09.2026", "26.09.2026", "27.09.2026",
    # Октябрь (добавлены только неохваченные)
    "10.10.2026", "11.10.2026", "17.10.2026", "18.10.2026", "24.10.2026", "25.10.2026",
    # Ноябрь (добавлены только неохваченные)
    "14.11.2026", "15.11.2026", "21.11.2026", "22.11.2026", "28.11.2026", "29.11.2026",
    # Декабрь
    "05.12.2026", "06.12.2026", "12.12.2026", "13.12.2026", "19.12.2026", "20.12.2026", "26.12.2026", "27.12.2026",
    # Канун Нового 2027 года
    "31.12.2026"
]
home_dir = os.path.expanduser("~")
directory = os.path.join(home_dir, "Downloads")
# Загружаем переменные из .env
load_dotenv()
login_NG = os.getenv("login_NG")
password_NG = os.getenv("password_NG")


def choosing_day(excluded_date):
    today = datetime.now().date()
    user_input = today
    days_count = 8
    excluded_date = [datetime.strptime(date, "%d.%m.%Y").date() for date in
                     excluded_date]  # делаем даты удобными для прочтения, к одному формату
    # основной цикл для нахождения даты
    while days_count != 0:
        if user_input in excluded_date:
            user_input += timedelta(days=1)
        else:
            user_input += timedelta(days=1)
            days_count -= 1
    print(user_input)
    return user_input


# Просроки Наш Город(НГ)--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def choosing_time_NG():
    timenow = pd.Timestamp(datetime.now()).strftime('%H-%M')
    return timenow


def process_ng_prosroki_file(timenow, filepath, excluded_dates):
    user_input = choosing_day(excluded_dates)
    df = pd.read_excel(filepath)

    # Очистка секунд в датах
    df['Регламентный срок у сообщения (Портал)'] = df['Регламентный срок у сообщения (Портал)'].apply(
        lambda x: x.replace(second=0))

    # Фильтр по дате ввода
    df = df[df['Регламентный срок у сообщения (Портал)'] <= pd.to_datetime(user_input)]
    today = datetime.now()

    # --- ИЗМЕНЕНИЕ НАЧАЛО: Маппинг применяется ДО разделения на префекта и основной df ---
    responsible_mapping = {
        'ГБУ «Автомобильные дороги ЮВАО»': 'АВД ЮВАО',
        'ГБУ Жилищник Выхино района Выхино-Жулебино города Москвы': 'Выхино-Жулебино',
        'Управа Выхино-Жулебино': 'Выхино-Жулебино',
        'ГБУ Жилищник Нижегородского района города Москвы': 'Нижегородский',
        'Управа Нижегородский': 'Нижегородский',
        'ГБУ Жилищник района Капотня города Москвы': 'Капотня',
        'Управа Капотня': 'Капотня',
        'ГБУ Жилищник района Кузьминки города Москвы': 'Кузьминки',
        'Управа Кузьминки': 'Кузьминки',
        'ГБУ Жилищник района Лефортово города Москвы': 'Лефортово',
        'Управа Лефортово': 'Лефортово',
        'ГБУ Жилищник района Люблино города Москвы': 'Люблино',
        'Управа Люблино': 'Люблино',
        'ГБУ Жилищник района Марьино города Москвы': 'Марьино',
        'Управа Марьино': 'Марьино',
        'ГБУ Жилищник района Некрасовка города Москвы': 'Некрасовка',
        'Управа Некрасовка': 'Некрасовка',
        'ГБУ Жилищник района Печатники города Москвы': 'Печатники',
        'Управа Печатники': 'Печатники',
        'ГБУ Жилищник района Текстильщики города Москвы': 'Текстильщики',
        'Управа Текстильщики': 'Текстильщики',
        'ГБУ Жилищник Рязанского района города Москвы': 'Рязанский',
        'Управа Рязанский': 'Рязанский',
        'ГБУ Жилищник Южнопортового района города Москвы': 'Южнопортовый',
        'Управа Южнопортовый': 'Южнопортовый'
    }

    # Создаем столбец 'Район' для ВСЕХ записей сразу по маппингу
    df['Район'] = df['Ответственный ОИВ первого уровня'].map(responsible_mapping)
    # --- ИЗМЕНЕНИЕ КОНЕЦ ---

    # условие для выделения просрочек ЛК ПРЕФЕКТА
    # (Фильтруем по столбцу Ответственный, но столбец Район уже создан по маппингу)
    condition = (df['Ответственный за подготовку ответа'].str.contains('Префектура Юго-Восточного округа')) & (
            df['Регламентный срок у сообщения (Портал)'] < today)
    prefect = df[condition].copy()  # .copy() чтобы избежать SettingWithCopyWarning

    # cоздаем сводную таблицу для префекта просрочек
    # Теперь индекс 'Район' точно существует и заполнен согласно маппингу
    pivot_prefect = pd.pivot_table(prefect, values='Номер заявки', index='Район', aggfunc='count')
    pivot_prefect = pivot_prefect.rename(columns={'Номер заявки': 'Кабинет префекта просрочки'})

    if pivot_prefect.empty:
        pivot_prefect = pd.DataFrame(
            index=['АВД ЮВАО', 'Выхино-Жулебино', 'Капотня', "Кузьминки", "Лефортово", 'Люблино', 'Марьино',
                   'Некрасовка', 'Нижегородский', 'Печатники', 'Рязанский', 'Текстильщики', 'Южнопортовый']
            , columns=['Кабинет префекта просрочки'])
    print(pivot_prefect)

    # выбрасываем просрочки префекта, а также все, что связанно с перефектурой за датафрейм
    # (Основной df очищается от строк префектуры для дальнейшей обработки)
    df = df[~df['Ответственный за подготовку ответа'].str.contains('Префектура Юго-Восточного округа')]

    # (Словарь responsible_mapping теперь не нужен здесь, так как применен выше)

    # устанавливаем формат даты
    excluded_dates_with_time = [
        datetime.strptime(date_str, "%d.%m.%Y").replace(hour=23, minute=59, second=0)
        for date_str in excluded_dates
    ]
    excluded_dates_dt = pd.to_datetime(excluded_dates_with_time)
    excluded_date = [datetime.strptime(date, "%d.%m.%Y").date() for date in excluded_dates]

    main_df = df.copy()

    def change_status(df):
        df = df.copy()
        df.loc[:, "Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "Готовится ответ", "Готовится ответ (ОИВ взял доп. срок)")
        df.loc[:, "Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "На доработке", "На доработке (Город вернул)")
        df.loc[:, "Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "На модерации", "На модерации (Проверка города)")
        df.loc[:, "Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "На утверждении", "На утверждении (У куратора)")
        df.loc[:, "Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "Нет ответа", "Нет ответа (ОИВ не дал ответ)")
        return df

    def table_is_none(date, number):
        df_empty = pd.DataFrame(
            index=['АВД ЮВАО', 'Выхино-Жулебино', 'Капотня', "Кузьминки", "Лефортово", 'Люблино', 'Марьино',
                   'Некрасовка', 'Нижегородский', 'Печатники', 'Рязанский', 'Текстильщики', 'Южнопортовый']
            , columns=[f'{number} день ({date.strftime('%d.%m')})']).fillna(0)
        print(f"{number}-й день пустой")
        return df_empty

    def crearing_day_in_svod(df, date, number):
        new_date = date + timedelta(days=1)
        while new_date in excluded_date:
            new_date += timedelta(days=1)
        # Проверка на пустой df перед операцией .dt.date
        if df.empty:
            return table_is_none(new_date, number), new_date

        df_date = change_status(df[df['Регламентный срок у сообщения (Портал)'].dt.date == new_date])
        pivot_date_for_svod = pd.pivot_table(df_date, values='Номер заявки', index='Район', aggfunc='count')
        new_name = f'{number} день ({new_date.strftime('%d.%m')})'
        if not pivot_date_for_svod.empty:
            pivot_date_for_svod.rename(columns={pivot_date_for_svod.columns[-1]: new_name}, inplace=True)
            return pivot_date_for_svod, new_date
        else:
            pivot_date_for_svod = table_is_none(new_date, number)
        return pivot_date_for_svod, new_date

    # 8-й день
    today_date = datetime.now().date()
    day_8 = today_date
    while day_8 in excluded_date:
        day_8 += timedelta(days=1)

    # Проверка на пустой main_df
    if not main_df.empty:
        df_date_8 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_8])
    else:
        df_date_8 = pd.DataFrame(columns=main_df.columns)

    pivot8_dlya_svoda = pd.pivot_table(df_date_8, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'{8} день ({day_8.strftime('%d.%m')})'
    if not pivot8_dlya_svoda.empty:
        pivot8_dlya_svoda.rename(columns={pivot8_dlya_svoda.columns[-1]: new_name}, inplace=True)

    pivot_8 = pd.pivot_table(df_date_8, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)

    if not pivot_8.empty:
        new_name = 'Всего'
        pivot_8.rename(columns={pivot_8.columns[-1]: new_name}, inplace=True)
        pivot_8.rename(index={pivot_8.index[-1]: new_name}, inplace=True)
    else:
        # Если pivot_8 пустой, pivot8_dlya_svoda мог быть обработан выше, но на всякий случай
        if pivot8_dlya_svoda.empty:
            pivot8_dlya_svoda = table_is_none(day_8, 8)

    # 7-й день
    day_7 = day_8 + timedelta(days=1)
    while day_7 in excluded_date:
        day_7 += timedelta(days=1)

    if not main_df.empty:
        df_date_7 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_7])
    else:
        df_date_7 = pd.DataFrame(columns=main_df.columns)

    pivot_7 = pd.pivot_table(df_date_7, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)

    pivot7_dlya_svoda = pd.pivot_table(df_date_7, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'{7} день ({day_7.strftime('%d.%m')})'
    if not pivot7_dlya_svoda.empty:
        pivot7_dlya_svoda.rename(columns={pivot7_dlya_svoda.columns[-1]: new_name}, inplace=True)
    if not pivot_7.empty:
        new_name = 'Всего'
        pivot_7.rename(columns={pivot_7.columns[-1]: new_name}, inplace=True)
        pivot_7.rename(index={pivot_7.index[-1]: new_name}, inplace=True)
    else:
        if pivot7_dlya_svoda.empty:
            pivot7_dlya_svoda = table_is_none(day_7, 7)

    # 6-й день
    day_6 = day_7 + timedelta(days=1)
    while day_6 in excluded_date:
        day_6 += timedelta(days=1)

    if not main_df.empty:
        df_date_6 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_6])
    else:
        df_date_6 = pd.DataFrame(columns=main_df.columns)

    pivot_6 = pd.pivot_table(df_date_6, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)

    pivot6_dlya_svoda = pd.pivot_table(df_date_6, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'{6} день ({day_6.strftime('%d.%m')})'
    if not pivot6_dlya_svoda.empty:
        pivot6_dlya_svoda.rename(columns={pivot6_dlya_svoda.columns[-1]: new_name}, inplace=True)
    if not pivot_6.empty:
        new_name = 'Всего'
        pivot_6.rename(columns={pivot_6.columns[-1]: new_name}, inplace=True)
        pivot_6.rename(index={pivot_6.index[-1]: new_name}, inplace=True)
    else:
        if pivot6_dlya_svoda.empty:
            pivot6_dlya_svoda = table_is_none(day_6, 6)

    # 5-й день
    day_5 = day_6 + timedelta(days=1)
    while day_5 in excluded_date:
        day_5 += timedelta(days=1)

    if not main_df.empty:
        df_date_5 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_5])
    else:
        df_date_5 = pd.DataFrame(columns=main_df.columns)

    pivot_5 = pd.pivot_table(df_date_5, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)

    pivot5_dlya_svoda = pd.pivot_table(df_date_5, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'{5} день ({day_5.strftime('%d.%m')})'
    if not pivot5_dlya_svoda.empty:
        pivot5_dlya_svoda.rename(columns={pivot5_dlya_svoda.columns[-1]: new_name}, inplace=True)
    if not pivot_5.empty:
        new_name = 'Всего'
        pivot_5.rename(columns={pivot_5.columns[-1]: new_name}, inplace=True)
        pivot_5.rename(index={pivot_5.index[-1]: new_name}, inplace=True)
    else:
        if pivot5_dlya_svoda.empty:
            pivot5_dlya_svoda = table_is_none(day_5, 5)

    # остальные дни
    # Передаем main_df, который уже имеет столбец 'Район'
    pivot4_dlya_svoda, date4 = crearing_day_in_svod(main_df, day_5, 4)
    pivot3_dlya_svoda, date3 = crearing_day_in_svod(main_df, date4, 3)
    pivot2_dlya_svoda, date2 = crearing_day_in_svod(main_df, date3, 2)
    pivot1_dlya_svoda, date1 = crearing_day_in_svod(main_df, date2, 1)

    # таблицы для просрочек
    if not main_df.empty:
        prosrok = main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date < today_date]
    else:
        prosrok = pd.DataFrame(columns=main_df.columns)

    prosrok_for_svod = pd.pivot_table(prosrok, values='Номер заявки', index='Район', aggfunc='count')
    prosrok_for_svod = prosrok_for_svod.rename(columns={'Номер заявки': 'Просрочки'})
    if prosrok_for_svod.empty:
        prosrok_for_svod = pd.DataFrame(
            index=['АВД ЮВАО', 'Выхино-Жулебино', 'Капотня', "Кузьминки", "Лефортово", 'Люблино', 'Марьино',
                   'Некрасовка', 'Нижегородский', 'Печатники', 'Рязанский', 'Текстильщики', 'Южнопортовый']
            , columns=['Просрочки']).fillna(0)

    df_prosrok = change_status(prosrok)
    if not df_prosrok.empty:
        pivot_prosrok = pd.pivot_table(df_prosrok, values='Номер заявки', index='Район',
                                       columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    else:
        pivot_prosrok = pd.DataFrame()

    if not pivot_prosrok.empty:
        new_name = 'Всего'
        pivot_prosrok.rename(columns={pivot_prosrok.columns[-1]: new_name}, inplace=True)
        pivot_prosrok.rename(index={pivot_prosrok.index[-1]: new_name}, inplace=True)
    else:
        print("Просроки пустые")

    # датафрейм для выходных
    holidays_df = main_df[main_df['Регламентный срок у сообщения (Портал)'].isin(excluded_dates_dt)]
    # датафрейм для выгрузки ответов в работе
    main_df = main_df[~main_df['Регламентный срок у сообщения (Портал)'].isin(excluded_dates_dt)].sort_values(
        by='Регламентный срок у сообщения (Портал)')

    dfs = [prosrok_for_svod, pivot8_dlya_svoda, pivot7_dlya_svoda, pivot6_dlya_svoda, pivot5_dlya_svoda,
           pivot4_dlya_svoda, pivot3_dlya_svoda, pivot2_dlya_svoda, pivot1_dlya_svoda]

    merged_df = reduce(lambda left, right: pd.merge(left, right, left_index=True, right_index=True, how='outer'), dfs)
    merged_table = pd.merge(pivot_prefect, merged_df, left_index=True, right_index=True, how='outer').fillna(0)
    all_in_work = pd.DataFrame({'Всего в работе': merged_table.sum(axis=1)}).fillna(0)
    all_urgent = pd.DataFrame({'Всего срочных': merged_table.iloc[:, :6].sum(axis=1)}).fillna(0)

    final_svod = pd.merge(all_in_work, pivot_prefect, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = pd.merge(final_svod, all_urgent, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = pd.merge(final_svod, merged_df, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = final_svod.sort_values(by='Всего срочных', ascending=False)

    totals_row = final_svod.sum(axis=0)
    totals_row.name = 'Итог по округу'
    df_totals = pd.DataFrame(totals_row).T
    df_with_totals = pd.concat([final_svod, df_totals])
    df_with_totals.index.name = 'Ответственный за подготовку ответа'

    # сохраняем по пути и добавляем листы
    # Убедитесь, что переменная directory определена в глобальной области видимости
    processed_file_path = os.path.join(directory,
                                       f"Ответы в работе_{datetime.now().strftime('%d.%m')}_на_{timenow}.xlsx")

    with pd.ExcelWriter(processed_file_path, engine='openpyxl') as writer:
        df_with_totals.to_excel(writer, sheet_name='СВОД', index=True, startrow=2)
        pivot_prosrok.to_excel(writer, sheet_name='просрочки', index=True, startrow=2)
        pivot_8.to_excel(writer, sheet_name='8-й день', index=True, startrow=2)
        pivot_7.to_excel(writer, sheet_name='7-й день', index=True, startrow=2)
        pivot_6.to_excel(writer, sheet_name='6-й день', index=True, startrow=2)
        pivot_5.to_excel(writer, sheet_name='5-й день', index=True, startrow=2)
        main_df.to_excel(writer, sheet_name='Ответы в работе', index=False, startrow=0)
        holidays_df.to_excel(writer, sheet_name='Выходные', index=False, startrow=0)
        # Теперь в листе Префект просрок столбец Район соответствует маппингу
        prefect.to_excel(writer, sheet_name='Префект просрок', index=False, startrow=0)

    return processed_file_path


async def parcing_data(context, chat_id):
    chrome_install = ChromeDriverManager().install()
    folder = os.path.dirname(chrome_install)
    chromedriver_path = os.path.join(folder, "chromedriver.exe")
    driver = webdriver.Chrome(service=ChromeService(chromedriver_path))
    try:
        # Откройте страницу логина
        driver.get('https://gorod.mos.ru/api/service/auth/auth')

        # Найдите поля для ввода логина и пароля и заполните их
        username = driver.find_element(By.XPATH, '//input[@placeholder="Логин *"]')
        password = driver.find_element(By.XPATH, '//input[@placeholder="Пароль*"]')
        username.send_keys(login_NG)
        password.send_keys(password_NG)

        # Найдите и нажмите кнопку логина
        login_button = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div[2]/form[1]/button')
        login_button.click()
        # Подождите, пока страница загрузится
        WebDriverWait(driver, 200).until(EC.presence_of_element_located((By.XPATH,
                                                                         '//div[@class="dashboard__block-link"]//div[@class="button-big link"]//div[@class="dashboard-container__links-title" and contains(text(), "Аналитика")]')))
        # переход в ответы в работе
        driver.get('https://gorod.mos.ru/admin/ker/olap/report/155')
        time.sleep(7)
        # # прыжок в меню
        # button = driver.find_element(By.XPATH,
        #                              "/html/body/div[3]/div/div[2]/div/div/div/div/form/header/div[1]/button[1]/span[2]/i")
        # button.click()
        # time.sleep(4)
        # # выбор фильтра
        # WebDriverWait(driver, 20).until(EC.presence_of_element_located(
        #     (By.XPATH, '/html/body/div[3]/div/div[2]/div/div/div/div/form/div[1]/aside/div/div[2]/div/div[1]/div/a')))
        # button = driver.find_element(By.XPATH,
        #                              "/html/body/div[3]/div/div[2]/div/div/div/div/form/div[1]/aside/div/div[2]/div/div[1]/div/a")
        # button.click()

        # экспорт
        WebDriverWait(driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span')))
        button = driver.find_element(By.XPATH,
                                     '/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span')
        button.click()
        time.sleep(1)
        # # ок- выгркзка с экселя
        # button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[2]/div/div[3]/button[2]/span[2]/span')
        # button.click()
        # time.sleep(1)

        # one more time click to export
        button = driver.find_element(By.XPATH, "//button[contains(@class, 'bg-primary')]//span[text()='Экспорт']")
        button.click()
        time.sleep(1)

        # переход в загрузки
        driver.get('https://gorod.mos.ru/admin/ker/olap/downloads')
        # Подождите, пока страница загрузится)
        WebDriverWait(driver, 1500).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')))
        # скачивание файла
        button = driver.find_element(By.XPATH,
                                     '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')
        button.click()
        time.sleep(15)
        return True
    except Exception as e:
        error_message = f"❌Произошла ошибка при выгрузке Ответы в работе(НГ). Пожалуйста, попробуйте еще раз."
        print(error_message)  # Выводим ошибку в консоль
        await context.bot.send_message(chat_id=chat_id, text=error_message)  # Отправляем сообщение в Telegram
        return False
    finally:
        driver.quit()


def personalizating_table_osn(timenow):
    # Получение пути к файлу на рабочем столе
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)

    # Выбор первого листа
    ws = wb.worksheets[0]

    first_table_range = 'A3:M17'
    header_range = 'A3:M3'  # Диапазон заголовков
    data_range = 'A4:M16'  # Диапазон данных (исключая последнюю строку)
    last_range = 'A17:M17'
    # Определение стилей
    light_blue_fill = PatternFill(start_color="5286d1", end_color="5286d1", fill_type="solid")
    pale_blue_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc", fill_type="solid")
    pink_fill = PatternFill(start_color="f7867e", end_color="f7867e", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    times_new_roman_font = Font(name='Times New Roman', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    start_row = 3
    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column
    # шапка большой таблицы
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range1 = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range1)
        ws[f'A{start_row - 1}'] = (
            f'Сводная информация по нарушениям сроков подготовки ответов на сообщения, поступившие на '
            f'централизованный портал "Наш город" по состоянию на {timenow} {datetime.now().strftime("%d.%m.%y")} г.')

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border
    ws.row_dimensions[2].height = 37
    # Добавляем черные границы ко всему диапазону
    thin = Side(border_style="thin", color="000000")  # Черная граница
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws['A2:M2']:
        for cell in row:
            cell.border = border
    # Применение стиля к заголовкам (первая строка)
    for cell in ws[header_range][0]:
        cell.fill = light_blue_fill
        cell.font = Font(name='Times New Roman', bold=True, size=11)
        cell.border = thin_border
        cell.alignment = center_alignment

    for cell in ws[last_range][0]:
        cell.font = Font(name='Times New Roman', bold=True, size=11)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Применение стиля к первому столбцу и следующим трем столбцам (A, B, C, D)
    for row in ws[data_range]:
        for cell in row[1:4]:  # Индексы 0, 1, 2, 3 соответствуют столбцам A, B, C, D
            cell.fill = pale_blue_fill
            cell.font = Font(name='Times New Roman', bold=False, size=11)
            cell.border = thin_border
            cell.alignment = center_alignment

    # Применение стиля ко всем значениям в таблице (делаем жирными)
    for row in ws[data_range]:
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name='Times New Roman', size=11)
            cell.alignment = center_alignment
        # Применение стиля к первому столбцу и следующим трем столбцам (A, B, C, D)
    for row in ws[data_range]:
        for cell in row[:1]:  # Индексы 0, 1, 2, 3 соответствуют столбцам A, B, C, D
            cell.fill = pale_blue_fill
            cell.font = Font(name='Times New Roman', bold=True, size=11)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for row in ws[data_range]:
        for cell in row[2:5]:  # Индексы 0, 1, 2, 3 соответствуют столбцам A, B, C, D
            cell.font = Font(name='Times New Roman', bold=True, size=11)
            cell.border = thin_border
            cell.alignment = center_alignment
    for row in ws[data_range]:
        for cell in row[3:4]:  # Индексы 0, 1, 2, 3 соответствуют столбцам A, B, C, D
            cell.font = Font(name='Times New Roman', bold=True, size=11, color="800000")
            cell.border = thin_border
            cell.alignment = center_alignment

    # Применение условного форматирования к указанным столбцам по индексу
    columns_to_format = [3, 5, 6, 7, 8, 9]  # Индексы столбцов (1-индексированные)
    for col_idx in columns_to_format:
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        col_range = f"{col_letter}4:{col_letter}16"  # Исключаем последнюю строку
        rule = CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=pink_fill)
        ws.conditional_formatting.add(col_range, rule)
        cell.alignment = center_alignment
    for row_num in range(4, 17):
        for col_num in range(10, ws.max_column + 1):  # Цикл по всем столбцам в строке
            ws.cell(row=row_num, column=col_num).fill = pale_blue_fill

    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(name='Times New Roman', size=11, bold=True)

    # Применение выравнивания и шрифта к шапке таблицы (например, строка 1)
    for cell in ws[3]:
        cell.alignment = header_alignment
        cell.font = header_font
        cell.alignment = center_alignment
    # ширина
    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 7.5
    ws.column_dimensions['G'].width = 7.5
    ws.column_dimensions['H'].width = 7.5
    ws.column_dimensions['I'].width = 7.5
    ws.column_dimensions['J'].width = 7.5
    ws.column_dimensions['K'].width = 7.5
    ws.column_dimensions['L'].width = 7.5
    ws.column_dimensions['M'].width = 7.5

    # высота
    ws.row_dimensions[3].height = 55
    ws.row_dimensions[4].height = 14.5
    ws.row_dimensions[5].height = 14.5
    ws.row_dimensions[6].height = 14.5
    ws.row_dimensions[7].height = 14.5
    ws.row_dimensions[8].height = 14.5
    ws.row_dimensions[9].height = 14.5
    ws.row_dimensions[10].height = 14.5
    ws.row_dimensions[11].height = 14.5
    ws.row_dimensions[12].height = 14.5
    ws.row_dimensions[13].height = 14.5
    ws.row_dimensions[14].height = 14.5
    ws.row_dimensions[15].height = 14.5
    ws.row_dimensions[16].height = 14.5
    ws.row_dimensions[17].height = 14.5
    # Сохранение изменений в файл
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_prosrok(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[1]  # Индексация начинается с 0, поэтому 1 - это второй лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = 'Просроченные сообщения в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_eight_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[2]  # Индексация начинается с 0, поэтому 1 - это второй лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '8-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_seven_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[3]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '7-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_six_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[4]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '6-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_five_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[5]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '5-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the five table in the first sheet in {file_path} successfully.')


def add_run_delete_and_save_files(timenow):
    # Открываем Excel через COM
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True  # Отображаем Excel

    # Открываем книгу
    wb = excel.Workbooks.Open(os.path.abspath(file_path))

    # Код макроса в виде строки
    vba_code = """  
Sub CopyTablesToFirstSheet()  
            Dim wsFirst As Worksheet  
            Dim wsSecond As Worksheet  
            Dim wsThird As Worksheet  
            Dim wsFour As Worksheet 
            Dim wsFive As Worksheet 
            Dim wsSix As Worksheet
            Dim lastRow As Long  
            Dim copyRange As Range  
            ' Установите ссылки на листы  
            Set wsFirst = ThisWorkbook.Worksheets(1)  ' Первый лист  
            Set wsSecond = ThisWorkbook.Worksheets(2) ' 2 лист  
            Set wsThird = ThisWorkbook.Worksheets(3)  ' 3 лист  
            Set wsFour = ThisWorkbook.Worksheets(4)  ' 4 лист 
            Set wsFive = ThisWorkbook.Worksheets(5)
            Set wsSix = ThisWorkbook.Worksheets(6)
            ' Копирование из второго листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Находим последнюю строку в первом листе  
            Set copyRange = wsSecond.UsedRange ' Выберите диапазон, который хотите скопировать  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист  
            ' Копирование из третьего листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Обновляем переменную lastRow  
            Set copyRange = wsThird.UsedRange ' Повторяем для третьего листа  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист  
            ' Копирование из четвертого листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Обновляем переменную lastRow  
            Set copyRange = wsFour.UsedRange ' Повторяем для третьего листа  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист  
            ' Копирование из пятого листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Обновляем переменную lastRow  
            Set copyRange = wsFive.UsedRange ' Повторяем для третьего листа  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист
            ' Копирование из шестого листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Обновляем переменную lastRow  
            Set copyRange = wsSix.UsedRange ' Повторяем для третьего листа  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист
            ' Очистка буфера обмена  
            Application.CutCopyMode = False  
            ' Удаление второго и третьего листов  
            Application.DisplayAlerts = False ' Отключаем предупреждения  
            wsSecond.Delete  
            wsThird.Delete  
            wsFour.Delete
            wsFive.Delete
            wsSix.Delete
            Application.DisplayAlerts = True ' Включаем предупреждения обратно  
        End Sub  

        Sub DeleteFirstSheet()  
            Dim wsFirst As Worksheet  
            Set wsFirst = ThisWorkbook.Worksheets(1)  ' Первый лист  
            Application.DisplayAlerts = False  ' Отключаем предупреждения  
            wsFirst.Delete  ' Удаляем первый лист  
            Application.DisplayAlerts = True  ' Включаем предупреждения обратно  
        End Sub
    """

    # Добавляем модуль в книгу и вставляем код
    vba_module = wb.VBProject.VBComponents.Add(1)  # 1 - это тип модуля стандартный
    vba_module.Name = 'MyMacroModule'  # Имя модуля
    vba_module.CodeModule.AddFromString(vba_code)

    # Сохраняем книгу
    # wb.Save()

    # Выполнение макроса
    excel.Application.Run('MyMacroModule.CopyTablesToFirstSheet')

    # Сохранение первого листа как PDF с заданным именем
    pdf_file_name = f'{datetime.now().strftime("%d.%m")}_на_{timenow}.pdf'
    pdf_path = os.path.join(os.path.dirname(file_path), pdf_file_name)  # Формируем путь к PDF
    wsFirst = wb.Worksheets(1)  # Ссылка на первый лист

    # Настройки страницы для печати
    wsFirst.PageSetup.FitToPagesWide = 1  # Устанавливаем количество страниц по ширине
    wsFirst.PageSetup.FitToPagesTall = 1  # Устанавливаем количество страниц по высоте на 1
    wsFirst.PageSetup.Zoom = False  # Отключаем масштабирование

    # Обновляем отступы страницы для уменьшения размера PDF
    wsFirst.PageSetup.LeftMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.RightMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.TopMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.BottomMargin = excel.Application.CentimetersToPoints(0.5)
    wb.Save()
    try:
        # Убираем ошибку, если файл уже существует
        if os.path.exists(pdf_path):
            print(f"Файл {pdf_path} существует. Удаление...")
            os.remove(pdf_path)  # Удаляем файл, если он существует
            print("Файл успешно удален.")

        print(f"Сохранение файла в {pdf_path}...")
        wsFirst.ExportAsFixedFormat(0, pdf_path)  # 0 - это xlTypePDF
        print(f"PDF успешно создан: {pdf_path}")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")

    first_sheet_file_name = f'CВОД_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx'
    first_sheet_file_path = os.path.join(directory, first_sheet_file_name)
    print(first_sheet_file_path)

    wsFirst.Copy()  # Копируем первый лист
    print(1)
    wb_first_sheet = excel.ActiveWorkbook  # Получаем ссылку на новый файл

    try:
        # Убираем ошибку, если файл уже существует
        if os.path.exists(first_sheet_file_path):
            print(f"Файл {first_sheet_file_path} существует. Удаление...")
            os.remove(first_sheet_file_path)  # Удаляем файл, если он существует
            print("Файл успешно удален.")

        print(f"Сохранение файла в {first_sheet_file_path}...")
        wb_first_sheet.SaveAs(first_sheet_file_path, FileFormat=51)
        print("Файл успешно сохранен.")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
    finally:
        wb_first_sheet.Close()  # Закрываем новый файл
        print(first_sheet_file_path)
    # Удаление первого листа из основного файла через макрос
    excel.Application.Run('MyMacroModule.DeleteFirstSheet')

    # Авторазмер колонок на оставшихся листах
    for sheet in wb.Worksheets:
        sheet.Cells.EntireColumn.AutoFit()

    # Сохраняем и закрываем основной файл
    wb.Save()
    wb.Close()
    excel.Quit()
    return pdf_path, first_sheet_file_path, file_path
