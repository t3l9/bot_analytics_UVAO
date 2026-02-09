from selenium.webdriver import Keys
from telegram import Update, ReplyKeyboardMarkup, InputFile, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes, \
    ConversationHandler, CallbackQueryHandler, CallbackContext
import time
import pandas as pd
import os
from datetime import datetime, timedelta
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
import win32com.client
from functools import reduce
import pythoncom


excluded_dates = [
    # Ваши исходные даты 2025 года
    "20.12.2025", "21.12.2025", "27.12.2025", "28.12.2025", "31.12.2025",

    # Новогодние каникулы и Рождество (2026)
    "01.01.2026", "02.01.2026", "03.01.2026", "04.01.2026", "05.01.2026",
    "06.01.2026", "07.01.2026", "08.01.2026", "09.01.2026", "10.01.2026",
    "11.01.2026",

    # Февральские праздники (23 февраля выпадает на понедельник, добавляем только его)
    "23.02.2026",
    # Предшествующие выходные
    "21.02.2026", "22.02.2026",

    # Мартовские праздники (8 марта - воскресенье, выходной переносится на 9 марта)
    "07.03.2026", "08.03.2026", "09.03.2026",

    # Майские праздники (1-3 мая и 9-11 мая)
    "01.05.2026", "02.05.2026", "03.05.2026",
    "09.05.2026", "10.05.2026", "11.05.2026",

    # День России (12 июня - пятница, длинные выходные 13-14 июня)
    "12.06.2026", "13.06.2026", "14.06.2026",

    # Ноябрьские праздники (4 ноября - среда, отдельный выходной)
    "04.11.2026",
    # Ближайшие выходные
    "31.10.2026", "01.11.2026", "07.11.2026", "08.11.2026",

    # Стандартные выходные 2026 года (субботы и воскресенья, не попавшие в периоды выше)
    # Январь
    "17.01.2026", "18.01.2026", "24.01.2026", "25.01.2026", "31.01.2026",
    # Февраль
    "01.02.2026", "07.02.2026", "08.02.2026", "14.02.2026", "15.02.2026", "28.02.2026",
    # Март
    "01.03.2026", "14.03.2026", "15.03.2026", "21.03.2026", "22.03.2026", "28.03.2026", "29.03.2026",
    # Апрель
    "04.04.2026", "05.04.2026", "11.04.2026", "12.04.2026", "18.04.2026", "19.04.2026", "25.04.2026", "26.04.2026",
    # Май (добавлены только неохваченные)
    "16.05.2026", "17.05.2026", "23.05.2026", "24.05.2026", "30.05.2026", "31.05.2026",
    # Июнь (добавлены только неохваченные)
    "06.06.2026", "07.06.2026", "20.06.2026", "21.06.2026", "27.06.2026", "28.06.2026",
    # Июль
    "04.07.2026", "05.07.2026", "11.07.2026", "12.07.2026", "18.07.2026", "19.07.2026", "25.07.2026", "26.07.2026",
    # Август
    "01.08.2026", "02.08.2026", "08.08.2026", "09.08.2026", "15.08.2026", "16.08.2026", "22.08.2026", "23.08.2026", "29.08.2026", "30.08.2026",
    # Сентябрь
    "05.09.2026", "06.09.2026", "12.09.2026", "13.09.2026", "19.09.2026", "20.09.2026", "26.09.2026", "27.09.2026",
    # Октябрь (добавлены только неохваченные)
    "10.10.2026", "11.10.2026", "17.10.2026", "18.10.2026", "24.10.2026", "25.10.2026",
    # Ноябрь (добавлены только неохваченные)
    "14.11.2026", "15.11.2026", "21.11.2026", "22.11.2026", "28.11.2026", "29.11.2026",
    # Декабрь
    "05.12.2026", "06.12.2026", "12.12.2026", "13.12.2026", "19.12.2026", "20.12.2026", "26.12.2026", "27.12.2026",
    # Канун Нового 2027 года
    "31.12.2026"
]
home_dir = os.path.expanduser("~")
# Путь к папке загрузок
directory = os.path.join(home_dir, "Downloads")
login_MM = 'UlyanovaMN1@mos.ru'
password_MM = 'hSJhfjksjw122!d'
login_NG = 'ulyanova_250124'
password_NG = 'Budva1608!@#$%^'
def choosing_day(excluded_date):
    today = datetime.now().date()
    user_input = today
    days_count = 8
    excluded_date = [datetime.strptime(date, "%d.%m.%Y").date() for date in excluded_date] #делаем даты удобными для прочтения, к одному формату
    #основной цикл для нахождения даты
    while days_count !=0:
      if user_input in excluded_date:
        user_input += timedelta(days=1)
      else:
        user_input += timedelta(days=1)
        days_count -=1
    print(user_input)
    return user_input
def choosing_time_frame_MM():
    today = datetime.now()
    weekday = today.weekday()
    start_of_week = today - timedelta(days=weekday)
    end_of_week = start_of_week + timedelta(days=6)
    if weekday == 0:
        start_day = today - timedelta(days=1)  # на один день назад
        end_day = today
    elif weekday == 1:
        start_day = start_of_week + timedelta(days=(weekday - 2))
        end_day = today
    elif weekday == 2:
        start_day = start_of_week + timedelta(days=(weekday - 3))
        end_day = today
    elif weekday == 3:
        start_day = start_of_week + timedelta(days=(weekday - 4))
        end_day = today
    elif weekday == 4:
        start_day = start_of_week + timedelta(days=(weekday - 5))
        end_day = today
    elif weekday == 5:
        start_day = start_of_week + timedelta(days=(weekday - 6))
        end_day = today
    elif weekday == 6:
        start_day = start_of_week + timedelta(days=(weekday - 7))
        end_day = today
    start_date = start_day.strftime("%d%m%Y")
    start_date = start_date + "2100"
    end_date = end_day.strftime("%d%m%Y")
    end_date = end_date + "2100"
    return start_date, end_date

# Просроки Наш Город(НГ)--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def choosing_time_NG():
    timenow = pd.Timestamp(datetime.now()).strftime('%H-%M')
    return timenow
def process_ng_prosroki_file(timenow, filepath, excluded_dates):
    user_input = choosing_day(excluded_dates)
    df = pd.read_excel(filepath)
    df['Регламентный срок у сообщения (Портал)'] = df['Регламентный срок у сообщения (Портал)'].apply(lambda x: x.replace(second=0))
    df = df[df['Регламентный срок у сообщения (Портал)'] <= pd.to_datetime(user_input)]      #оставляем только те сообщения, котоыре меньше или равны заданной даты
    today = datetime.now()

    #условие для выделения просрочек ЛК ПРЕФЕКТА
    condition = (df['Ответственный за подготовку ответа'].str.contains('Префектура Юго-Восточного округа')) & (
            df['Регламентный срок у сообщения (Портал)'] < today)
    prefect = df[condition]
    #cоздаем сводную таблицу для префекта просрочек
    pivot_prefect = pd.pivot_table(prefect, values='Номер заявки', index='Район', aggfunc='count')
    pivot_prefect = pivot_prefect.rename(columns={'Номер заявки': 'Кабинет префекта просрочки'})
    if pivot_prefect.empty:
        pivot_prefect = pd.DataFrame(
            index=['АВД ЮВАО', 'Выхино-Жулебино', 'Капотня', "Кузьминки", "Лефортово", 'Люблино', 'Марьино',
                   'Некрасовка', 'Нижегородский', 'Печатники', 'Рязанский', 'Текстильщики', 'Южнопортовый']
            , columns=['Кабинет префекта просрочки'])
    print(pivot_prefect)

    #выбрасываем просрочки префекта, а также все, что связанно с перефектурой за датафрейм
    df = df[~df['Ответственный за подготовку ответа'].str.contains('Префектура Юго-Восточного округа')]
    responsible_mapping = {
        'ГБУ «Автомобильные дороги ЮВАО»': 'АВД ЮВАО',
        'ГБУ Жилищник Выхино района Выхино-Жулебино города Москвы': 'Выхино-Жулебино',
        'Управа Выхино-Жулебино': 'Выхино-Жулебино',
        'ГБУ Жилищник Нижегородского района города Москвы': 'Нижегородский',
        'Управа Нижегородский': 'Нижегородский',
        'ГБУ Жилищник района Капотня города Москвы': 'Капотня',
        'Управа Капотня': 'Капотня',
        'ГБУ Жилищник района Кузьминки города Москвы': 'Кузьминки',
        'Управа Кузьминки': 'Кузьминки',
        'ГБУ Жилищник района Лефортово города Москвы': 'Лефортово',
        'Управа Лефортово': 'Лефортово',
        'ГБУ Жилищник района Люблино города Москвы': 'Люблино',
        'Управа Люблино': 'Люблино',
        'ГБУ Жилищник района Марьино города Москвы': 'Марьино',
        'Управа Марьино': 'Марьино',
        'ГБУ Жилищник района Некрасовка города Москвы': 'Некрасовка',
        'Управа Некрасовка': 'Некрасовка',
        'ГБУ Жилищник района Печатники города Москвы': 'Печатники',
        'Управа Печатники': 'Печатники',
        'ГБУ Жилищник района Текстильщики города Москвы': 'Текстильщики',
        'Управа Текстильщики': 'Текстильщики',
        'ГБУ Жилищник Рязанского района города Москвы': 'Рязанский',
        'Управа Рязанский': 'Рязанский',
        'ГБУ Жилищник Южнопортового района города Москвы': 'Южнопортовый',
        'Управа Южнопортовый': 'Южнопортовый'
    }
    df['Район'] = df['Ответственный за подготовку ответа'].map(responsible_mapping)
    #устанавливаем формат даты
    excluded_dates_with_time = [
        datetime.strptime(date_str, "%d.%m.%Y").replace(hour=23, minute=59, second=0)
        for date_str in excluded_dates
    ]
    excluded_dates_dt = pd.to_datetime(excluded_dates_with_time)
    excluded_date = [datetime.strptime(date, "%d.%m.%Y").date() for date in excluded_dates]
    #df['Регламентный срок у сообщения (Портал)'] = pd.to_datetime(df['Регламентный срок у сообщения (Портал)'])
    main_df = df.copy()
    #Фильтруем датафрейм, исключая даты из excluded_dates
    #main_df['Регламентный срок у сообщения (Портал)'] = pd.to_datetime(main_df['Регламентный срок у сообщения (Портал)'])


    def change_status(df):
        df.loc[:,"Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "Готовится ответ", "Готовится ответ (ОИВ взял доп. срок)")
        df.loc[:,"Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "На доработке","На доработке (Город вернул)")
        df.loc[:,"Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "На модерации", "На модерации (Проверка города)")
        df.loc[:,"Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "На утверждении", "На утверждении (У куратора)")
        df.loc[:,"Статус подготовки ответа на сообщение"] = df["Статус подготовки ответа на сообщение"].replace(
            "Нет ответа", "Нет ответа (ОИВ не дал ответ)")
        return df
    def table_is_none(date, number):
        df = pd.DataFrame(
            index=['АВД ЮВАО', 'Выхино-Жулебино', 'Капотня', "Кузьминки", "Лефортово", 'Люблино', 'Марьино',
                   'Некрасовка', 'Нижегородский', 'Печатники', 'Рязанский', 'Текстильщики', 'Южнопортовый']
            , columns=[f'{number} день ({date.strftime('%d.%m')})']).fillna(0)
        print(f"{number}-й день пустой")
        return df
    def crearing_day_in_svod(df,date,number):
        new_date = date + timedelta(days=1)
        while new_date in excluded_date:
            new_date += timedelta(days=1)
        df_date = change_status(df[df['Регламентный срок у сообщения (Портал)'].dt.date == new_date])
        pivot_date_for_svod = pd.pivot_table(df_date, values='Номер заявки', index='Район',aggfunc='count')
        new_name = f'{number} день ({new_date.strftime('%d.%m')})'
        if not pivot_date_for_svod.empty:
            pivot_date_for_svod.rename(columns={pivot_date_for_svod.columns[-1]: new_name}, inplace=True)
            return pivot_date_for_svod, new_date
        else:
            pivot_date_for_svod = table_is_none(new_date, number)
        return pivot_date_for_svod, new_date
    #8-й день
    today = datetime.now().date()
    day_8 = today
    #если дата в выходных, то идти дальше, присваивая 8-му дню след дату
    while day_8 in excluded_date:
        day_8 += timedelta(days=1)
    df_date_8 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_8])
    pivot8_dlya_svoda = pd.pivot_table(df_date_8, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'{8} день ({day_8.strftime('%d.%m')})'
    if not pivot8_dlya_svoda.empty:
        pivot8_dlya_svoda.rename(columns={pivot8_dlya_svoda.columns[-1]: new_name}, inplace=True)
    pivot_8 = pd.pivot_table(df_date_8, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    #pivot_8 = pivot_8.fillna(0).astype("int")
    if not pivot_8.empty:
        new_name = 'Всего'
        pivot_8.rename(columns={pivot_8.columns[-1]: new_name}, inplace=True)
        # Замена названия последней строки
        pivot_8.rename(index={pivot_8.index[-1]: new_name}, inplace=True)
    else:
        pivot8_dlya_svoda = table_is_none(day_8, 8)


    #7-й день
    day_7 = day_8 + timedelta(days = 1)
    while day_7 in excluded_date:
        day_7 +=timedelta(days=1)
    df_date_7 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_7])
    pivot_7 = pd.pivot_table(df_date_7, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    #pivot_7 = pivot_7.fillna(0).astype("int")

    pivot7_dlya_svoda = pd.pivot_table(df_date_7, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'{7} день ({day_7.strftime('%d.%m')})'
    if not pivot7_dlya_svoda.empty:
        pivot7_dlya_svoda.rename(columns={pivot7_dlya_svoda.columns[-1]: new_name}, inplace=True)
    if not pivot_7.empty:
        new_name = 'Всего'
        pivot_7.rename(columns={pivot_7.columns[-1]: new_name}, inplace=True)
        # Замена названия последней строки
        pivot_7.rename(index={pivot_7.index[-1]: new_name}, inplace=True)
    else:
        pivot7_dlya_svoda = table_is_none(day_7, 7)

    #6-й день
    day_6 = day_7 + timedelta(days = 1)
    while day_6 in excluded_date:
        day_6 +=timedelta(days=1)
    df_date_6 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_6])
    pivot_6 = pd.pivot_table(df_date_6, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)

    pivot6_dlya_svoda = pd.pivot_table(df_date_6, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'{6} день ({day_6.strftime('%d.%m')})'
    if not pivot6_dlya_svoda.empty:
        pivot6_dlya_svoda.rename(columns={pivot6_dlya_svoda.columns[-1]: new_name}, inplace=True)
    if not pivot_6.empty:
        new_name = 'Всего'
        pivot_6.rename(columns={pivot_6.columns[-1]: new_name}, inplace=True)
        # Замена названия последней строки
        pivot_6.rename(index={pivot_6.index[-1]: new_name}, inplace=True)
    else:
        pivot6_dlya_svoda = table_is_none(day_6, 6)

    #5-й день
    day_5 = day_6 + timedelta(days = 1)
    while day_5 in excluded_date:
        day_5 +=timedelta(days=1)
    df_date_5 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_5])
    pivot_5 = pd.pivot_table(df_date_5, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)

    pivot5_dlya_svoda = pd.pivot_table(df_date_5, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'{5} день ({day_5.strftime('%d.%m')})'
    if not pivot5_dlya_svoda.empty:
        pivot5_dlya_svoda.rename(columns={pivot5_dlya_svoda.columns[-1]: new_name}, inplace=True)
    if not pivot_5.empty:
        new_name = 'Всего'
        pivot_5.rename(columns={pivot_5.columns[-1]: new_name}, inplace=True)
        # Замена названия последней строки
        pivot_5.rename(index={pivot_5.index[-1]: new_name}, inplace=True)
    else:
        pivot5_dlya_svoda = table_is_none(day_5, 5)
    #остальные дни
    pivot4_dlya_svoda, date4 = crearing_day_in_svod(main_df, day_5, 4)
    pivot3_dlya_svoda, date3 = crearing_day_in_svod(main_df, date4, 3)
    pivot2_dlya_svoda, date2 = crearing_day_in_svod(main_df, date3, 2)
    pivot1_dlya_svoda, date1 = crearing_day_in_svod(main_df, date2, 1)
    #таблицы для просрочек
    prosrok = main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date < today]
    prosrok_for_svod = pd.pivot_table(prosrok, values='Номер заявки', index='Район', aggfunc='count')
    prosrok_for_svod = prosrok_for_svod.rename(columns={'Номер заявки': 'Просрочки'})
    if prosrok_for_svod.empty:
        prosrok_for_svod = pd.DataFrame(
            index=['АВД ЮВАО', 'Выхино-Жулебино', 'Капотня', "Кузьминки", "Лефортово", 'Люблино', 'Марьино',
                   'Некрасовка', 'Нижегородский', 'Печатники', 'Рязанский', 'Текстильщики', 'Южнопортовый']
            , columns=['Просрочки']).fillna(0)
    df_prosrok = change_status(prosrok)
    pivot_prosrok = pd.pivot_table(df_prosrok, values='Номер заявки', index='Район',
                                   columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    #pivot_prosrok.fillna(0).astype("int")
    # замена all
    if not pivot_prosrok.empty:
        new_name = 'Всего'
        pivot_prosrok.rename(columns={pivot_prosrok.columns[-1]: new_name}, inplace=True)
        # Замена названия последней строки
        pivot_prosrok.rename(index={pivot_prosrok.index[-1]: new_name}, inplace=True)
    else:
        print("Просроки пустые")


    #датафрейм для выходных
    holidays_df = main_df[main_df['Регламентный срок у сообщения (Портал)'].isin(excluded_dates_dt)]
    #датафрейм для выгрузки ответов в работе
    main_df = main_df[~main_df['Регламентный срок у сообщения (Портал)'].isin(excluded_dates_dt)].sort_values(by='Регламентный срок у сообщения (Портал)')

    dfs = [prosrok_for_svod,  pivot8_dlya_svoda, pivot7_dlya_svoda, pivot6_dlya_svoda, pivot5_dlya_svoda, pivot4_dlya_svoda, pivot3_dlya_svoda,pivot2_dlya_svoda, pivot1_dlya_svoda]
    # Объединение всех датафреймов по ключу
    #merged_df = reduce(lambda left, right: pd.merge(left, right, on = "Район", how='outer'), dfs)
    merged_df = reduce(lambda left, right: pd.merge(left, right, left_index=True, right_index=True, how='outer'), dfs)
    merged_table = pd.merge(pivot_prefect, merged_df, left_index=True, right_index=True, how='outer').fillna(0)
    all_in_work = pd.DataFrame({'Всего в работе': merged_table.sum(axis=1)}).fillna(0) #столбец всего в работе
    all_urgent = pd.DataFrame({'Всего срочных': merged_table.iloc[:,:6].sum(axis=1)}).fillna(0)
    #мерджим все в финальную таблицу
    final_svod = pd.merge(all_in_work, pivot_prefect, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = pd.merge(final_svod, all_urgent, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = pd.merge(final_svod, merged_df, left_index=True, right_index=True, how='outer').fillna(0)
    #сортируем по столбцу со срочными
    final_svod = final_svod.sort_values(by='Всего срочных', ascending=False)
    #добавляем итог
    totals_row = final_svod.sum(axis=0)
    totals_row.name = 'Итог по округу'
    df_totals = pd.DataFrame(totals_row).T
    df_with_totals = pd.concat([final_svod, df_totals])
    # дополнительно переименовывем и сохраняем с наванием и нужной датой
    df_with_totals.index.name = 'Ответственный за подготовку ответа'

    #сохраняем по пути и добавляем листы
    processed_file_path = os.path.join(directory,
                                       f"Ответы в работе_{datetime.now().strftime('%d.%m')}_на_{timenow}.xlsx")
    df.to_excel(processed_file_path, index=False)
    # cохраняем файлы
    with pd.ExcelWriter(processed_file_path, engine='openpyxl') as writer:
        df_with_totals.to_excel(writer, sheet_name='СВОД', index=True, startrow=2)
        pivot_prosrok.to_excel(writer, sheet_name='просрочки', index=True, startrow=2)
        pivot_8.to_excel(writer, sheet_name='8-й день', index=True, startrow=2)
        pivot_7.to_excel(writer, sheet_name='7-й день', index=True, startrow=2)
        pivot_6.to_excel(writer, sheet_name='6-й день', index=True, startrow=2)
        pivot_5.to_excel(writer, sheet_name='5-й день', index=True, startrow=2)
        main_df.to_excel(writer, sheet_name='Ответы в работе', index=False, startrow=0)
        holidays_df.to_excel(writer, sheet_name='Выходные', index=False, startrow=0)
        prefect.to_excel(writer, sheet_name='Префект просрок', index=False, startrow=0)
    return processed_file_path
async def parcing_data(context, chat_id):
    chrome_install = ChromeDriverManager().install()
    folder = os.path.dirname(chrome_install)
    chromedriver_path = os.path.join(folder, "chromedriver.exe")
    driver = webdriver.Chrome(service=ChromeService(chromedriver_path))
    try:
        # Откройте страницу логина
        driver.get('https://gorod.mos.ru/api/service/auth/auth')

        # Найдите поля для ввода логина и пароля и заполните их
        username = driver.find_element(By.XPATH, '//input[@placeholder="Логин *"]')
        password = driver.find_element(By.XPATH, '//input[@placeholder="Пароль*"]')
        username.send_keys(login_NG)
        password.send_keys(password_NG)

        # Найдите и нажмите кнопку логина
        login_button = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div[2]/form[1]/button')
        login_button.click()
        # Подождите, пока страница загрузится
        WebDriverWait(driver, 200).until(EC.presence_of_element_located((By.XPATH,
                                                                        '//div[@class="dashboard__block-link"]//div[@class="button-big link"]//div[@class="dashboard-container__links-title" and contains(text(), "Аналитика")]')))
        # переход в ответы в работе
        driver.get('https://gorod.mos.ru/admin/ker/olap/report/155')
        time.sleep(7)
        # # прыжок в меню
        # button = driver.find_element(By.XPATH,
        #                              "/html/body/div[3]/div/div[2]/div/div/div/div/form/header/div[1]/button[1]/span[2]/i")
        # button.click()
        # time.sleep(4)
        # # выбор фильтра
        # WebDriverWait(driver, 20).until(EC.presence_of_element_located(
        #     (By.XPATH, '/html/body/div[3]/div/div[2]/div/div/div/div/form/div[1]/aside/div/div[2]/div/div[1]/div/a')))
        # button = driver.find_element(By.XPATH,
        #                              "/html/body/div[3]/div/div[2]/div/div/div/div/form/div[1]/aside/div/div[2]/div/div[1]/div/a")
        # button.click()

        # экспорт
        WebDriverWait(driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span')))
        button = driver.find_element(By.XPATH,
                                     '/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span')
        button.click()
        time.sleep(1)
        # # ок- выгркзка с экселя
        # button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[2]/div/div[3]/button[2]/span[2]/span')
        # button.click()
        # time.sleep(1)

        #one more time click to export
        button = driver.find_element(By.XPATH, "//button[contains(@class, 'bg-primary')]//span[text()='Экспорт']")
        button.click()
        time.sleep(1)

        # переход в загрузки
        driver.get('https://gorod.mos.ru/admin/ker/olap/downloads')
        # Подождите, пока страница загрузится)
        WebDriverWait(driver, 1500).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')))
        # скачивание файла
        button = driver.find_element(By.XPATH,
                                     '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')
        button.click()
        time.sleep(15)
        return True
    except Exception as e:
        error_message = f"❌Произошла ошибка при выгрузке Ответы в работе(НГ). Пожалуйста, попробуйте еще раз."
        print(error_message)  # Выводим ошибку в консоль
        await context.bot.send_message(chat_id=chat_id, text=error_message)  # Отправляем сообщение в Telegram
        await context.bot.send_message(chat_id=chat_id, text='Выберите команду:',
                                       reply_markup=InlineKeyboardMarkup([
                                           [InlineKeyboardButton("🏢 ЛК префекта(НГ)", callback_data='lk_prefekt')],
                                           [InlineKeyboardButton("📊 Монитор в Работе(ММ)", callback_data='mm_monitor')],
                                           [InlineKeyboardButton("📈 Ответы в работе (НГ)", callback_data='ng_answers')],
                                           [InlineKeyboardButton("📋 СВОД МЖИ(НГ)", callback_data='mji_summary')],
                                           [InlineKeyboardButton("📅 Статистика МЖИ", callback_data='mji_stat')],
                                           [InlineKeyboardButton("🌐 Монитор (НГ)", callback_data='city_monitor')],
                                           [InlineKeyboardButton("♻️ КП/БП (ММ)", callback_data='mm_kp_bp')],
                                           [InlineKeyboardButton("🚨 Ежедневные просрочки (ММ)", callback_data='today_mm')],
                                           [InlineKeyboardButton("🔄 Последнее обновление",
                                                                 callback_data='last_update')],
                                           [InlineKeyboardButton("❓ Объяснение команд", callback_data='explain')],
                                       ]))
        return False
    finally:
        driver.quit()
def personalizating_table_osn(timenow):
    # Получение пути к файлу на рабочем столе
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)

    # Выбор первого листа
    ws = wb.worksheets[0]

    first_table_range = 'A3:M17'
    header_range = 'A3:M3'  # Диапазон заголовков
    data_range = 'A4:M16'  # Диапазон данных (исключая последнюю строку)
    last_range = 'A17:M17'
    # Определение стилей
    light_blue_fill = PatternFill(start_color="5286d1", end_color="5286d1", fill_type="solid")
    pale_blue_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc", fill_type="solid")
    pink_fill = PatternFill(start_color="f7867e", end_color="f7867e", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    times_new_roman_font = Font(name='Times New Roman', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    start_row = 3
    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column
    # шапка большой таблицы
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range1 = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range1)
        ws[f'A{start_row - 1}'] = (
            f'Сводная информация по нарушениям сроков подготовки ответов на сообщения, поступившие на '
            f'централизованный портал "Наш город" по состоянию на {timenow} {datetime.now().strftime("%d.%m.%y")} г.')

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border
    ws.row_dimensions[2].height = 37
    # Добавляем черные границы ко всему диапазону
    thin = Side(border_style="thin", color="000000")  # Черная граница
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws['A2:M2']:
        for cell in row:
            cell.border = border
    # Применение стиля к заголовкам (первая строка)
    for cell in ws[header_range][0]:
        cell.fill = light_blue_fill
        cell.font = Font(name='Times New Roman', bold=True, size=11)
        cell.border = thin_border
        cell.alignment = center_alignment

    for cell in ws[last_range][0]:
        cell.font = Font(name='Times New Roman', bold=True, size=11)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Применение стиля к первому столбцу и следующим трем столбцам (A, B, C, D)
    for row in ws[data_range]:
        for cell in row[1:4]:  # Индексы 0, 1, 2, 3 соответствуют столбцам A, B, C, D
            cell.fill = pale_blue_fill
            cell.font = Font(name='Times New Roman', bold=False, size=11)
            cell.border = thin_border
            cell.alignment = center_alignment

    # Применение стиля ко всем значениям в таблице (делаем жирными)
    for row in ws[data_range]:
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name='Times New Roman', size=11)
            cell.alignment = center_alignment
        # Применение стиля к первому столбцу и следующим трем столбцам (A, B, C, D)
    for row in ws[data_range]:
        for cell in row[:1]:  # Индексы 0, 1, 2, 3 соответствуют столбцам A, B, C, D
            cell.fill = pale_blue_fill
            cell.font = Font(name='Times New Roman', bold=True, size=11)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for row in ws[data_range]:
        for cell in row[2:5]:  # Индексы 0, 1, 2, 3 соответствуют столбцам A, B, C, D
            cell.font = Font(name='Times New Roman', bold=True, size=11)
            cell.border = thin_border
            cell.alignment = center_alignment
    for row in ws[data_range]:
        for cell in row[3:4]:  # Индексы 0, 1, 2, 3 соответствуют столбцам A, B, C, D
            cell.font = Font(name='Times New Roman', bold=True, size=11, color="800000")
            cell.border = thin_border
            cell.alignment = center_alignment

    # Применение условного форматирования к указанным столбцам по индексу
    columns_to_format = [3, 5, 6, 7, 8, 9]  # Индексы столбцов (1-индексированные)
    for col_idx in columns_to_format:
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        col_range = f"{col_letter}4:{col_letter}16"  # Исключаем последнюю строку
        rule = CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=pink_fill)
        ws.conditional_formatting.add(col_range, rule)
        cell.alignment = center_alignment
    for row_num in range(4, 17):
        for col_num in range(10, ws.max_column + 1):  # Цикл по всем столбцам в строке
            ws.cell(row=row_num, column=col_num).fill = pale_blue_fill

    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(name='Times New Roman', size=11, bold=True)

    # Применение выравнивания и шрифта к шапке таблицы (например, строка 1)
    for cell in ws[3]:
        cell.alignment = header_alignment
        cell.font = header_font
        cell.alignment = center_alignment
    # ширина
    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 7.5
    ws.column_dimensions['G'].width = 7.5
    ws.column_dimensions['H'].width = 7.5
    ws.column_dimensions['I'].width = 7.5
    ws.column_dimensions['J'].width = 7.5
    ws.column_dimensions['K'].width = 7.5
    ws.column_dimensions['L'].width = 7.5
    ws.column_dimensions['M'].width = 7.5

    # высота
    ws.row_dimensions[3].height = 55
    ws.row_dimensions[4].height = 14.5
    ws.row_dimensions[5].height = 14.5
    ws.row_dimensions[6].height = 14.5
    ws.row_dimensions[7].height = 14.5
    ws.row_dimensions[8].height = 14.5
    ws.row_dimensions[9].height = 14.5
    ws.row_dimensions[10].height = 14.5
    ws.row_dimensions[11].height = 14.5
    ws.row_dimensions[12].height = 14.5
    ws.row_dimensions[13].height = 14.5
    ws.row_dimensions[14].height = 14.5
    ws.row_dimensions[15].height = 14.5
    ws.row_dimensions[16].height = 14.5
    ws.row_dimensions[17].height = 14.5
    # Сохранение изменений в файл
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')
def personalizating_table_prosrok(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[1]  # Индексация начинается с 0, поэтому 1 - это второй лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = 'Просроченные сообщения в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')
def personalizating_table_eight_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[2]  # Индексация начинается с 0, поэтому 1 - это второй лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '8-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')
def personalizating_table_seven_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[3]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '7-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')
def personalizating_table_six_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[4]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '6-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')
def personalizating_table_five_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[5]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '5-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the five table in the first sheet in {file_path} successfully.')
def add_run_delete_and_save_files(timenow):
    # Открываем Excel через COM
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True  # Отображаем Excel

    # Открываем книгу
    wb = excel.Workbooks.Open(os.path.abspath(file_path))

    # Код макроса в виде строки
    vba_code = """  
Sub CopyTablesToFirstSheet()  
            Dim wsFirst As Worksheet  
            Dim wsSecond As Worksheet  
            Dim wsThird As Worksheet  
            Dim wsFour As Worksheet 
            Dim wsFive As Worksheet 
            Dim wsSix As Worksheet
            Dim lastRow As Long  
            Dim copyRange As Range  
            ' Установите ссылки на листы  
            Set wsFirst = ThisWorkbook.Worksheets(1)  ' Первый лист  
            Set wsSecond = ThisWorkbook.Worksheets(2) ' 2 лист  
            Set wsThird = ThisWorkbook.Worksheets(3)  ' 3 лист  
            Set wsFour = ThisWorkbook.Worksheets(4)  ' 4 лист 
            Set wsFive = ThisWorkbook.Worksheets(5)
            Set wsSix = ThisWorkbook.Worksheets(6)
            ' Копирование из второго листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Находим последнюю строку в первом листе  
            Set copyRange = wsSecond.UsedRange ' Выберите диапазон, который хотите скопировать  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист  
            ' Копирование из третьего листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Обновляем переменную lastRow  
            Set copyRange = wsThird.UsedRange ' Повторяем для третьего листа  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист  
            ' Копирование из четвертого листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Обновляем переменную lastRow  
            Set copyRange = wsFour.UsedRange ' Повторяем для третьего листа  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист  
            ' Копирование из пятого листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Обновляем переменную lastRow  
            Set copyRange = wsFive.UsedRange ' Повторяем для третьего листа  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист
            ' Копирование из шестого листа  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2 ' Обновляем переменную lastRow  
            Set copyRange = wsSix.UsedRange ' Повторяем для третьего листа  
            copyRange.Copy wsFirst.Cells(lastRow, 1) ' Вставляем скопированный диапазон в первый лист
            ' Очистка буфера обмена  
            Application.CutCopyMode = False  
            ' Удаление второго и третьего листов  
            Application.DisplayAlerts = False ' Отключаем предупреждения  
            wsSecond.Delete  
            wsThird.Delete  
            wsFour.Delete
            wsFive.Delete
            wsSix.Delete
            Application.DisplayAlerts = True ' Включаем предупреждения обратно  
        End Sub  

        Sub DeleteFirstSheet()  
            Dim wsFirst As Worksheet  
            Set wsFirst = ThisWorkbook.Worksheets(1)  ' Первый лист  
            Application.DisplayAlerts = False  ' Отключаем предупреждения  
            wsFirst.Delete  ' Удаляем первый лист  
            Application.DisplayAlerts = True  ' Включаем предупреждения обратно  
        End Sub
    """

    # Добавляем модуль в книгу и вставляем код
    vba_module = wb.VBProject.VBComponents.Add(1)  # 1 - это тип модуля стандартный
    vba_module.Name = 'MyMacroModule'  # Имя модуля
    vba_module.CodeModule.AddFromString(vba_code)

    # Сохраняем книгу
    #wb.Save()

    # Выполнение макроса
    excel.Application.Run('MyMacroModule.CopyTablesToFirstSheet')

    # Сохранение первого листа как PDF с заданным именем
    pdf_file_name = f'{datetime.now().strftime("%d.%m")}_на_{timenow}.pdf'
    pdf_path = os.path.join(os.path.dirname(file_path), pdf_file_name)  # Формируем путь к PDF
    wsFirst = wb.Worksheets(1)  # Ссылка на первый лист

    # Настройки страницы для печати
    wsFirst.PageSetup.FitToPagesWide = 1  # Устанавливаем количество страниц по ширине
    wsFirst.PageSetup.FitToPagesTall = 1  # Устанавливаем количество страниц по высоте на 1
    wsFirst.PageSetup.Zoom = False  # Отключаем масштабирование

    # Обновляем отступы страницы для уменьшения размера PDF
    wsFirst.PageSetup.LeftMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.RightMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.TopMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.BottomMargin = excel.Application.CentimetersToPoints(0.5)
    wb.Save()
    try:
        # Убираем ошибку, если файл уже существует
        if os.path.exists(pdf_path):
            print(f"Файл {pdf_path} существует. Удаление...")
            os.remove(pdf_path)  # Удаляем файл, если он существует
            print("Файл успешно удален.")

        print(f"Сохранение файла в {pdf_path}...")
        wsFirst.ExportAsFixedFormat(0, pdf_path)  # 0 - это xlTypePDF
        print(f"PDF успешно создан: {pdf_path}")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")

    first_sheet_file_name = f'CВОД_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx'
    first_sheet_file_path = os.path.join(directory, first_sheet_file_name)
    print(first_sheet_file_path)

    wsFirst.Copy()  # Копируем первый лист
    print(1)
    wb_first_sheet = excel.ActiveWorkbook  # Получаем ссылку на новый файл

    try:
        # Убираем ошибку, если файл уже существует
        if os.path.exists(first_sheet_file_path):
            print(f"Файл {first_sheet_file_path} существует. Удаление...")
            os.remove(first_sheet_file_path)  # Удаляем файл, если он существует
            print("Файл успешно удален.")

        print(f"Сохранение файла в {first_sheet_file_path}...")
        wb_first_sheet.SaveAs(first_sheet_file_path, FileFormat=51)
        print("Файл успешно сохранен.")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
    finally:
        wb_first_sheet.Close()  # Закрываем новый файл
        print(first_sheet_file_path)
    # Удаление первого листа из основного файла через макрос
    excel.Application.Run('MyMacroModule.DeleteFirstSheet')

    # Авторазмер колонок на оставшихся листах
    for sheet in wb.Worksheets:
        sheet.Cells.EntireColumn.AutoFit()

    # Сохраняем и закрываем основной файл
    wb.Save()
    wb.Close()
    excel.Quit()
    return pdf_path, first_sheet_file_path, file_path



#Монитор ММ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
async def parcing_data_MM(context, chat_id, MM_start_date, MM_end_date):
    chrome_install = ChromeDriverManager().install()
    folder = os.path.dirname(chrome_install)
    chromedriver_path = os.path.join(folder, "chromedriver.exe")
    driver = webdriver.Chrome(service=ChromeService(chromedriver_path))
    driver.maximize_window()
    try:
        # Откройте страницу логина
        driver.get('https://arm-mmonitor.mos.ru')
        time.sleep(0.5)
        # Найдите поля для ввода логина и пароля и заполните их
        username = driver.find_element(By.XPATH, '/html/body/main/div/div[2]/div/form[1]/div[1]/div/input')
        password = driver.find_element(By.XPATH, '/html/body/main/div/div[2]/div/form[1]/div[2]/div/input')
        username.send_keys(login_MM)
        password.send_keys(password_MM)
        # Найдите и нажмите кнопку логина
        login_button = driver.find_element(By.XPATH, '/html/body/main/div/div[2]/div/form[1]/div[5]/div[1]/button')
        login_button.click()

        # Подождите, пока страница загрузится

        WebDriverWait(driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/section/section/main/div/div[1]/div[2]/span[1]')))
        time.sleep(0.3)
        button = driver.find_element(By.XPATH, "/html/body/div[1]/div/section/section/main/div/div[1]/div[2]/span[1]")
        button.click()

        # выпадающая дата
        button = driver.find_element(By.XPATH,
                                     "/html/body/div[2]/div/div[2]/div/div/div[2]/div[1]/label[4]/span/div/div")
        button.click()
        time.sleep(0.5)
        # ставим дата отчета
        button = driver.find_element(By.XPATH,
                                     "/html/body/div[2]/div/div[2]/div/div/div[2]/div[1]/label[4]/span/div/div/div[2]/div/div/div/div[2]/div[1]/div/div/div[3]/div")
        button.click()
        time.sleep(1)
        # enter start date
        button1 = driver.find_element(By.XPATH,
                                      '/html/body/div[2]/div/div[2]/div/div/div[2]/div[1]/label[4]/div/div[1]/div/input')
        button1.click()
        button1.send_keys(Keys.CONTROL + 'a')  # Выделить весь текст
        button1.send_keys(Keys.BACKSPACE)  # Удалить выделенный текст
        time.sleep(0.3)
        button1.send_keys(MM_start_date)  # дата начала вводится
        time.sleep(0.5)
        # enter end date
        button2 = driver.find_element(By.XPATH,
                                      '/html/body/div[2]/div/div[2]/div/div/div[2]/div[1]/label[4]/div/div[2]/div/input')
        button2.click()
        button2.send_keys(Keys.CONTROL + 'a')  # Выделить весь текст
        button2.send_keys(Keys.BACKSPACE)  # Удалить выделенный текст
        time.sleep(0.3)
        button2.send_keys(MM_end_date)  # дата конца вводится

        # доходим до ответсвенных
        button = driver.find_element(By.XPATH,
                                     '/html/body/div[2]/div/div[2]/div/div/div[2]/div[1]/label[13]/div/div[1]/div')
        button.click()
        time.sleep(0.5)
        # выбираем территориальные
        button = driver.find_element(By.XPATH,
                                     '/html/body/div[2]/div/div[2]/div/div/div[2]/div[1]/label[13]/div/div[2]/div/div/div/div[2]/div[1]/div/div/div[2]/div')
        button.click()
        time.sleep(0.5)
        # нажимаем показать
        button = driver.find_element(By.XPATH,
                                     '/html/body/div[2]/div/div[2]/div/div/div[2]/div[2]/button[1]')
        button.click()
        # time.sleep(1000)
        time.sleep(0.5)
        # нажимаем на кнопку что бы закрыть фильтры
        body = driver.find_element(By.TAG_NAME, 'body')
        body.click()
        time.sleep(0.5)
        # добавляем в очередь скачивания
        button = driver.find_element(By.CSS_SELECTOR, "svg.icon.xls-icon")
        button.click()
        time.sleep(0.5)
        # переходим в загрузки
        driver.get('https://arm-mmonitor.mos.ru/#/export-files')
        # обновляем страницу пока не появится нужный элемент, затем скачиваем
        i = 0
        while i < 50:
            try:
                # Ожидание элемента в течение 5 секунд (без обновления страницы)
                element = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH,
                                                                                           "/html/body/div/div/section/section/main/div/div/div[1]/div/div/div/div/div[2]/table/tbody/tr[3]/td[5]/div/button/span")))
                time.sleep(1)
                element.click()
                print("Элемент найден, прекращаем обновление страницы.")
                break  # Выход из цикла, если элемент найден
            except:
                print("Элемент не найден, обновляем страницу.")
                driver.refresh()  # Обновление страницы
                i += 1
                print(i)
                time.sleep(3)  # Ожидание 3 секунд перед следующей проверкой
        time.sleep(6)
        return True
    except Exception as e:
        error_message = f"❌Произошла ошибка при выгрузке ММ. Пожалуйста, попробуйте еще раз."
        print(error_message)  # Выводим ошибку в консоль
        await context.bot.send_message(chat_id=chat_id, text=error_message)  # Отправляем сообщение в Telegram
        await context.bot.send_message(chat_id=chat_id, text='Выберите команду:',
                                       reply_markup=InlineKeyboardMarkup([
                                           [InlineKeyboardButton("🏢 ЛК префекта(НГ)", callback_data='lk_prefekt')],
                                           [InlineKeyboardButton("📊 Монитор в Работе(ММ)", callback_data='mm_monitor')],
                                           [InlineKeyboardButton("📈 Ответы в работе (НГ)", callback_data='ng_answers')],
                                           [InlineKeyboardButton("📋 СВОД МЖИ(НГ)", callback_data='mji_summary')],
                                           [InlineKeyboardButton("📅 Статистика МЖИ", callback_data='mji_stat')],
                                           [InlineKeyboardButton("🌐 Монитор (НГ)", callback_data='city_monitor')],
                                           [InlineKeyboardButton("♻️ КП/БП (ММ)", callback_data='mm_kp_bp')],
                                           [InlineKeyboardButton("🚨 Ежедневные просрочки (ММ)", callback_data='today_mm')],
                                           [InlineKeyboardButton("🔄 Последнее обновление",
                                                                 callback_data='last_update')],
                                           [InlineKeyboardButton("❓ Объяснение команд", callback_data='explain')],
                                       ]))
        return False
    finally:
        driver.quit()
def choosing_time_MM():
    today = datetime.now()
    current_date = pd.Timestamp(datetime.now().date())

    eight_am_today = current_date + pd.Timedelta(hours=0)
    ten_am_today = current_date + pd.Timedelta(hours=10, minutes=59, seconds=59)

    twelf_am_today = current_date + pd.Timedelta(hours=11)
    therteen_am_today = current_date + pd.Timedelta(hours=14, minutes=59, seconds=59)

    three_pm_today = current_date + pd.Timedelta(hours=15)
    five_am_today = current_date + pd.Timedelta(hours=19, minutes=59, seconds=59)

    eight_pm_today = current_date + pd.Timedelta(hours=20)
    eleven_pm_today = current_date + pd.Timedelta(hours=23, minutes=59, seconds=59)

    if (today > eight_am_today) & (today < ten_am_today):
        timenow = "УТРО"
    elif (today > twelf_am_today) & (today < therteen_am_today):
        timenow = "ДЕНЬ"
    elif (today > three_pm_today) & (today < five_am_today):
        timenow = "ВЕЧЕР"
    elif (today > eight_pm_today) & (today < eleven_pm_today):
        timenow = "НОЧЬ"
    return timenow
def first_attribute(df):
    today = datetime.now()
    weekday = today.weekday()
    # Определяем начало и конец текущей недели
    start_of_week = today - timedelta(days=weekday)
    end_of_week = start_of_week + timedelta(days=6)
    # Фильтруем DataFrame в соответствии с требуемой логикой
    if weekday == 0:
        df.loc[(df['Просрок'] == 'Да') & (df['Статус в системе'] == 'Устранено')
           & (df[
                  'Срок устранения до'].dt.date == today.date()), "ТипСПросроком"] = "Устранено с нарушением срока " + today.strftime(
        "%d.%m.%y") + " (На текущей уб. неделе)"
    elif weekday == 1:
        start_day = start_of_week + timedelta(days=(weekday - 1))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df[
                    'Статус в системе'] == 'Устранено'), "ТипСПросроком"] = "Устранено с нарушением срока " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (На текущей уб. неделе)"
    elif weekday == 2:
        start_day = start_of_week + timedelta(days=(weekday - 2))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df[
                    'Статус в системе'] == 'Устранено'), "ТипСПросроком"] = "Устранено с нарушением срока " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (На текущей уб. неделе)"
    elif weekday == 3:
        start_day = start_of_week + timedelta(days=(weekday - 3))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df[
                    'Статус в системе'] == 'Устранено'), "ТипСПросроком"] = "Устранено с нарушением срока " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (На текущей уб. неделе)"
    elif weekday == 4:
        start_day = start_of_week + timedelta(days=(weekday - 4))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df[
                    'Статус в системе'] == 'Устранено'), "ТипСПросроком"] = "Устранено с нарушением срока " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (На текущей уб. неделе)"
    elif weekday == 5:
        start_day = start_of_week + timedelta(days=(weekday - 5))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df[
                    'Статус в системе'] == 'Устранено'), "ТипСПросроком"] = "Устранено с нарушением срока " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (На текущей уб. неделе)"
    elif weekday == 6:
        start_day = today - timedelta(days=6)
        end_day = today
        df.loc[(df['Просрок'] == 'Да') & (df['Статус в системе'] == 'Устранено') & (df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()), "ТипСПросроком"] = ("Устранено с нарушением срока " + start_day.strftime("%d.%m.%y") + " по " + today.strftime("%d.%m.%y")) + " (На текущей уб. неделе)"
def second_attribute(df):
    today = datetime.now()
    weekday = today.weekday()
    # Определяем начало и конец текущей недели
    start_of_week = today - timedelta(days=weekday)
    if weekday == 6:
        start_day = today - timedelta(days=6)
        end_day = today
        df.loc[(df['Просрок'] == 'Да') & (df['Статус в системе'] == 'В работе') & (
                df['Срок устранения до'].dt.date >= start_day.date())
               & (df['Срок устранения до'].dt.date <= end_day.date()), "ТипСПросроком"] = (
                                                                                                  "В работе с просроком " + start_day.strftime(
                                                                                              "%d.%m.%y") + " по " + today.strftime(
                                                                                              "%d.%m.%y")) + " (Текущая уб. неделя)"

    elif weekday == 0:
        df.loc[(df['Просрок'] == 'Да') & (df['Статус в системе'] == 'В работе')
               & (df[
                      'Срок устранения до'].dt.date == today.date()), "ТипСПросроком"] = "В работе с просроком " + today.strftime(
            "%d.%m.%y") + " (Текущая уб. неделя)"

    elif weekday == 1:
        start_day = start_of_week + timedelta(days=(weekday - 1))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df['Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (Текущая уб. неделя)"

    elif weekday == 2:
        start_day = start_of_week + timedelta(days=(weekday - 2))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df['Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (Текущая уб. неделя)"

    elif weekday == 3:
        start_day = start_of_week + timedelta(days=(weekday - 3))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df['Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (Текущая уб. неделя)"

    elif weekday == 4:
        start_day = start_of_week + timedelta(days=(weekday - 4))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df['Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (Текущая уб. неделя)"

    elif weekday == 5:
        start_day = start_of_week + timedelta(days=(weekday - 5))
        end_day = today
        df.loc[(df['Срок устранения до'].dt.date >= start_day.date()) &
               (df['Срок устранения до'].dt.date <= end_day.date()) & (df['Просрок'] == 'Да') &
               (df['Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком " + start_day.strftime(
            "%d.%m.%y") + " по " + today.strftime("%d.%m.%y") + " (Текущая уб. неделя)"
def third_attribute(df):
    today = datetime.now()
    weekday = today.weekday()
    # Определяем начало и конец текущей недели
    if weekday == 0:
        end_of_last_week = today - timedelta(days=1)
        start_of_last_week = end_of_last_week - timedelta(days=6)
        df.loc[(df['Срок устранения до'].dt.date >= start_of_last_week.date()) &
               (df['Срок устранения до'].dt.date <= end_of_last_week.date()) & (df['Просрок'] == 'Да') &
               (df[
                    'Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком с " + start_of_last_week.strftime(
            "%d.%m.%y") + " по " + end_of_last_week.strftime("%d.%m.%y") + " (Прошедшая уб. неделя)"
    else:
        end_of_last_week = today - timedelta(days=(weekday+1))
        start_of_last_week = end_of_last_week - timedelta(days=6)
        df.loc[(df['Срок устранения до'].dt.date >= start_of_last_week.date()) &
               (df['Срок устранения до'].dt.date <= end_of_last_week.date()) & (df['Просрок'] == 'Да') &
               (df[
                    'Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком с " + start_of_last_week.strftime(
            "%d.%m.%y") + " по " + end_of_last_week.strftime("%d.%m.%y") + " (Прошедшая уб. неделя)"
def fourth_attribute(df):
    today = datetime.now()
    weekday = today.weekday()
    # Определяем начало и конец текущей недели
    earliest_date = df['Срок устранения до'].min()
    # if weekday == 0:
    #     end_of_last_week = today - timedelta(days=7)
    #     end_of_last_week_mon = end_of_last_week - timedelta(days=7)
    #     df.loc[(df['Срок устранения до'].dt.date >= earliest_date.date()) &
    #            (df['Срок устранения до'].dt.date <= end_of_last_week_mon.date()) & (df['Просрок'] == 'Да') &
    #            (df['Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком с " + earliest_date.strftime(
    #         "%d.%m.%y") + " по " + end_of_last_week_mon.strftime("%d.%m.%y") + " (Старые)"
    if weekday == 0:
        end_of_last_week = today - timedelta(days=1)
        end_of_last_week_mon = end_of_last_week - timedelta(days=7)
        df.loc[(df['Срок устранения до'].dt.date >= earliest_date.date()) &
               (df['Срок устранения до'].dt.date <= end_of_last_week_mon.date()) & (df['Просрок'] == 'Да') &
               (df[
                    'Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком с " + earliest_date.strftime(
            "%d.%m.%y") + " по " + end_of_last_week_mon.strftime("%d.%m.%y") + " (Старые)"
    else:
        end_of_last_week = today - timedelta(days=(weekday+1))
        end_of_last_week_mon = end_of_last_week - timedelta(days=7)
        df.loc[(df['Срок устранения до'].dt.date >= earliest_date.date()) &
               (df['Срок устранения до'].dt.date <= end_of_last_week_mon.date()) & (df['Просрок'] == 'Да') &
               (df[
                    'Статус в системе'] == 'В работе'), "ТипСПросроком"] = "В работе с просроком с " + earliest_date.strftime(
            "%d.%m.%y") + " по " + end_of_last_week_mon.strftime("%d.%m.%y") + " (Старые)"
def fifth_attribute(df):
    today = datetime.now()
    if today:
        df.loc[(df['Срок устранения до'].dt.date == today.date()) & (df['Просрок'] == 'Нет') &
               (df['Статус в системе'] == 'В работе'), "ТипБезПросрока"] = "Срок с " + pd.Timestamp(
            datetime.now()).strftime('%H:%M') + " " + today.strftime("%d.%m.%y") + " (Сегодня)"
def sixth_attribute(df):
    today = datetime.now()
    tommorow = today + timedelta(days=1)
    max_date = df[(df['Просрок'] == 'Нет') &
                  (df['Статус в системе'] == 'В работе')]['Срок устранения до'].max()
    if today:
        df.loc[((df['Срок устранения до'].dt.date >= tommorow.date()) & (
                df['Срок устранения до'].dt.date <= max_date.date()) & (df['Просрок'] == 'Нет') &
                (df['Статус в системе'] == 'В работе')) |
               ((df['Обещание устранения'].dt.date >= tommorow.date()) & (
                       df['Обещание устранения'].dt.date <= max_date.date()) & (df['Просрок'] == 'Нет') &
                (df['Статус в системе'] == 'В работе')), "ТипБезПросрока"] = "Срок с " + tommorow.strftime(
            "%d.%m.%y") + " по " + max_date.strftime("%d.%m.%y")
def snow_today(df):
    today = datetime.now()
    if today:
        df.loc[(df['Дата фиксации нарушения'].dt.date == today.date()) &
               ((df['Проблема'] == 'Наличие снега, наледи') | (df['Проблема'] == 'Неочищенная кровля')), "ТипСнег"] = "Снег " + today.strftime("%d.%m.%y") + " (Сегодня)"
def snow_all_expect_today(df):
    today = datetime.now()
    tomorrow = today - timedelta(days=1)
    weekday = today.weekday()
    # Определяем начало и конец текущей недели
    start_of_week = today - timedelta(days=weekday)
    if weekday == 6:
        start_day = today - timedelta(days=6)
        end_day = tomorrow
        df.loc[((df['Проблема'] == 'Наличие снега, наледи') | (df['Проблема'] == 'Неочищенная кровля')) & (
                df['Дата фиксации нарушения'].dt.date >= start_day.date())
               & (df['Дата фиксации нарушения'].dt.date <= end_day.date()), "ТипСнег"] = "Снег с " + start_day.strftime(
            "%d.%m.%y") + " по " + tomorrow.strftime(
            "%d.%m.%y") + " (Текущая уб. неделя)"

    # во monday в этом столбце ничего не будет, т.к. данный снег будет находиться в другом столбце (снег сегодня)
    elif weekday == 1:
        start_day = start_of_week + timedelta(days=(weekday - 1))
        end_day = tomorrow
        df.loc[((df['Проблема'] == 'Наличие снега, наледи') | (df['Проблема'] == 'Неочищенная кровля')) & (
                df['Дата фиксации нарушения'].dt.date >= start_day.date())
               & (df['Дата фиксации нарушения'].dt.date <= end_day.date()), "ТипСнег"] = "Снег " + tomorrow.strftime(
            "%d.%m.%y") + " (Текущая уб. неделя)"

    elif weekday == 2:
        start_day = start_of_week + timedelta(days=(weekday - 2))
        end_day = tomorrow
        df.loc[((df['Проблема'] == 'Наличие снега, наледи') | (df['Проблема'] == 'Неочищенная кровля')) & (
                df['Дата фиксации нарушения'].dt.date >= start_day.date())
               & (df['Дата фиксации нарушения'].dt.date <= end_day.date()), "ТипСнег"] = "Снег с " + start_day.strftime(
            "%d.%m.%y") + " по " + tomorrow.strftime(
            "%d.%m.%y") + " (Текущая уб. неделя)"

    elif weekday == 3:
        start_day = start_of_week + timedelta(days=(weekday - 3))
        end_day = tomorrow
        df.loc[((df['Проблема'] == 'Наличие снега, наледи') | (df['Проблема'] == 'Неочищенная кровля')) & (
                df['Дата фиксации нарушения'].dt.date >= start_day.date())
               & (df['Дата фиксации нарушения'].dt.date <= end_day.date()), "ТипСнег"] = "Снег с " + start_day.strftime(
            "%d.%m.%y") + " по " + tomorrow.strftime(
            "%d.%m.%y") + " (Текущая уб. неделя)"

    elif weekday == 4:
        start_day = start_of_week + timedelta(days=(weekday - 4))
        end_day = tomorrow
        df.loc[((df['Проблема'] == 'Наличие снега, наледи') | (df['Проблема'] == 'Неочищенная кровля')) & (
                df['Дата фиксации нарушения'].dt.date >= start_day.date())
               & (df['Дата фиксации нарушения'].dt.date <= end_day.date()), "ТипСнег"] = "Снег с " + start_day.strftime(
            "%d.%m.%y") + " по " + tomorrow.strftime(
            "%d.%m.%y") + " (Текущая уб. неделя)"

    elif weekday == 5:
        start_day = start_of_week + timedelta(days=(weekday - 5))
        end_day = tomorrow
        df.loc[((df['Проблема'] == 'Наличие снега, наледи') | (df['Проблема'] == 'Неочищенная кровля')) & (
                df['Дата фиксации нарушения'].dt.date >= start_day.date())
               & (df['Дата фиксации нарушения'].dt.date <= end_day.date()), "ТипСнег"] = "Снег с " + start_day.strftime(
            "%d.%m.%y") + " по " + tomorrow.strftime(
            "%d.%m.%y") + " (Текущая уб. неделя)"
def process_file_MM(filepath, timenow):
    df = pd.read_excel(filepath)
    # Список значений, которые должны присутствовать в столбце "Балансодержатель"
    wanted_values = [
        'ГБУ «Автомобильные дороги ЮВАО»',
        'ГБУ «Жилищник Выхино района Выхино-Жулебино»',
        'ГБУ «Жилищник Нижегородского района»',
        'ГБУ «Жилищник района Капотня»',
        'ГБУ «Жилищник района Кузьминки»',
        'ГБУ «Жилищник района Лефортово»',
        'ГБУ «Жилищник района Люблино»',
        'ГБУ «Жилищник района Марьино»',
        'ГБУ «Жилищник района Некрасовка»',
        'ГБУ «Жилищник района Печатники»',
        'ГБУ «Жилищник района Текстильщики»',
        'ГБУ «Жилищник района Южнопортовый»',
        'ГБУ «Жилищник Рязанского района»'
    ]
    df = df[df['Балансодержатель'].isin(wanted_values)]

    responsible_mapping = {
        'ГБУ «Автомобильные дороги ЮВАО»': 'АВД ЮВАО',
        'ГБУ «Жилищник Выхино района Выхино-Жулебино»': 'Выхино-Жулебино',
        'Управа района Выхино-Жулебино': 'Выхино-Жулебино',
        'ГБУ «Жилищник Нижегородского района»': 'Нижегородский',
        'Управа Нижегородского района': 'Нижегородский',
        'ГБУ «Жилищник района Капотня»': 'Капотня',
        'Управа района Капотня': 'Капотня',
        'ГБУ «Жилищник района Кузьминки»': 'Кузьминки',
        'Управа района Кузьминки': 'Кузьминки',
        'ГБУ «Жилищник района Лефортово»': 'Лефортово',
        'Управа района Лефортово': 'Лефортово',
        'ГБУ «Жилищник района Люблино»': 'Люблино',
        'Управа района Люблино': 'Люблино',
        'ГБУ «Жилищник района Марьино»': 'Марьино',
        'Управа района Марьино': 'Марьино',
        'ГБУ «Жилищник района Некрасовка»': 'Некрасовка',
        'Управа района Некрасовка': 'Некрасовка',
        'ГБУ «Жилищник района Печатники»': 'Печатники',
        'Управа района Печатники': 'Печатники',
        'ГБУ «Жилищник района Текстильщики»': 'Текстильщики',
        'Управа района Текстильщики': 'Текстильщики',
        'ГБУ «Жилищник Рязанского района»': 'Рязанский',
        'Управа Рязанского района': 'Рязанский',
        'ГБУ «Жилищник района Южнопортовый»': 'Южнопортовый',
        'Управа Южнопортового района': 'Южнопортовый'
    }
    df['Район'] = df['Ответственный исполнитель'].map(responsible_mapping)

    df['Срок устранения до'] = pd.to_datetime(df['Срок устранения до'])
    df['Обещание устранения'] = pd.to_datetime(df['Обещание устранения'])
    df['ТипБезПросрока'] = ''
    df['ТипСПросроком'] = ''
    df['ТипСнег'] = ''
    first_attribute(df)
    second_attribute(df)
    third_attribute(df)
    fourth_attribute(df)
    fifth_attribute(df)
    sixth_attribute(df)
    print(df[df["Проблема"] == "Наличие снега, наледи"])
    if not df[df["Проблема"].isin(["Наличие снега, наледи", "Неочищенная кровля"])].empty:
        print("Есть снег")
        snow_today(df)
        snow_all_expect_today(df)
    processed_file_path = os.path.join(directory,
                                       f"Монитор в работе_{timenow}_{datetime.now().strftime('%d.%m.%y')}.xlsx")
    df.to_excel(processed_file_path, index=False)
    with pd.ExcelWriter(processed_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='СВОД', index=False, startrow=0)
    excel_file = processed_file_path
    # VBA код макроса, который будет добавлен в Excel
    vba_macro = """  
Sub CreatePivotTable1()
    Dim wsData As Worksheet
    Dim wsPivot As Worksheet
    Dim pivotCache As PivotCache
    Dim pivotTable As PivotTable
    Dim lastRow As Long
    Dim lastCol As Long
    Dim foundTodayColumn As Boolean
    Dim cell As Range

    ' Укажите лист с данными
    Set wsData = ThisWorkbook.Sheets("СВОД") ' Замените на имя вашего листа с данными

    ' Создаем новый лист для сводной таблицы
    On Error Resume Next
    Application.DisplayAlerts = False
    ThisWorkbook.Sheets("Сводная таблица").Delete ' Удаляем лист, если уже существует
    Application.DisplayAlerts = True
    On Error GoTo 0
    Set wsPivot = ThisWorkbook.Sheets.Add
    wsPivot.Name = "Сводная таблица"

    ' Находим последний заполненный ряд и столбец на листе с данными
    lastRow = wsData.Cells(wsData.Rows.Count, "A").End(xlUp).Row
    lastCol = wsData.Cells(1, wsData.Columns.Count).End(xlToLeft).Column

    ' Создаем кэш для сводной таблицы
    Set pivotCache = ThisWorkbook.PivotCaches.Create( _
        SourceType:=xlDatabase, _
        SourceData:=wsData.Cells(1, 1).Resize(lastRow, lastCol))

    ' Создаем сводную таблицу
    Set pivotTable = pivotCache.CreatePivotTable( _
        TableDestination:=wsPivot.Cells(3, 1), _
        TableName:="MyPivotTable")

    ' Настройка сводной таблицы: строки - "Район", столбцы - "ТипБезПросрока", значения - количество строк
    With pivotTable
        .PivotFields("Район").Orientation = xlRowField
        .PivotFields("ТипБезПросрока").Orientation = xlColumnField
        .AddDataField .PivotFields("ID нарушения"), "Количество", xlCount
    End With
    With pivotTable
        .GrandTotalName = "На устранении без просрока" ' Замените на нужное название для общего итога
    End With
    wsPivot.Range("A4").Value = "Район"
    ' Скрываем первую строку
    wsPivot.Rows(3).Hidden = True

    ' Убираем столбец "Пусто" (где ТипБезПросрока неопределен)
    Dim typePivotField As PivotField
    Set typePivotField = pivotTable.PivotFields("ТипБезПросрока")
    For Each item In typePivotField.PivotItems
        If item.Name = "(blank)" Then
            item.Visible = False
        End If
    Next item

    ' Обновляем сводную таблицу
    pivotTable.RefreshTable

    ' Форматирование сводной таблицы
    Dim rng As Range
    Set rng = wsPivot.Range("A4").CurrentRegion
    With rng
        .Font.Name = "Times New Roman"
        .Font.Size = 14
        .Font.Bold = True
        .Borders.LineStyle = xlContinuous
        .WrapText = True ' Перенос текста
        .HorizontalAlignment = xlCenter ' Выравнивание по центру
        .VerticalAlignment = xlCenter
    End With
    wsPivot.Columns("A").ColumnWidth = 24 ' Установите желаемую ширину столбца
    With rng
        .HorizontalAlignment = xlCenter ' Горизонтальное выравнивание по центру
        .VerticalAlignment = xlCenter ' Вертикальное выравнивание по центру
    End With

    ' Настройка высоты строк и ширины столбцов
    wsPivot.Range("6:16").RowHeight = 19
    wsPivot.Columns("B").ColumnWidth = 39
    wsPivot.Columns("C").ColumnWidth = 34
    wsPivot.Columns("D").ColumnWidth = 33
    wsPivot.Columns("E").ColumnWidth = 39

    ' Проверяем наличие столбца с названием, содержащим "Сегодня"
    foundTodayColumn = False
    For Each cell In wsPivot.Range("B4:E4")
        If InStr(1, cell.Value, "Сегодня", vbTextCompare) > 0 Then
            foundTodayColumn = True 
            cell.Font.Color = RGB(255, 0, 0) ' Красный цвет текста заголовка
            Dim dataRange As Range
            Dim lastDataRow As Long
            lastDataRow = wsPivot.Cells(wsPivot.Rows.Count, cell.Column).End(xlUp).Row - 1 ' Уменьшаем на одну строку для исключения итогов
            Set dataRange = wsPivot.Range(cell.Offset(1, 0), wsPivot.Cells(lastDataRow, cell.Column))
            ' Применяем заливку к значениям > 0, исключая итоги
            For Each dataCell In dataRange
                If IsNumeric(dataCell.Value) And dataCell.Value > 0 Then
                    dataCell.Interior.Color = RGB(247, 134, 126) ' Красная заливка для положительных значений
                End If
            Next dataCell
        End If
    Next cell
End Sub
"""
    vba_macro2 = """Sub CreatePivotTable2()
    Dim wsData As Worksheet
    Dim wsPivot As Worksheet
    Dim pivotCache As PivotCache
    Dim pivotTable As PivotTable
    Dim lastRow As Long
    Dim lastCol As Long
    Dim pivotStartRow As Long

    ' Укажите лист с данными
    Set wsData = ThisWorkbook.Sheets("СВОД") ' Замените на имя вашего листа с данными

    ' Укажите существующий лист для сводной таблицы
    Set wsPivot = ThisWorkbook.Sheets("Сводная таблица") ' Замените на имя вашего листа с существующей сводной таблицей

    ' Находим последний заполненный ряд и столбец на листе с данными
    lastRow = wsData.Cells(wsData.Rows.Count, "A").End(xlUp).Row
    lastCol = wsData.Cells(1, wsData.Columns.Count).End(xlToLeft).Column

    ' Находим строку, где уже существует сводная таблица, и добавляем новую через 3 строки
    pivotStartRow = wsPivot.Cells(wsPivot.Rows.Count, 1).End(xlUp).Row + 3

    ' Создаем кэш для сводной таблицы
    Set pivotCache = ThisWorkbook.PivotCaches.Create( _
        SourceType:=xlDatabase, _
        SourceData:=wsData.Cells(1, 1).Resize(lastRow, lastCol))

    ' Создаем сводную таблицу
    Set pivotTable = pivotCache.CreatePivotTable( _
        TableDestination:=wsPivot.Cells(pivotStartRow, 1), _
        TableName:="MyPivotTableWithExpiration")

    ' Настройка сводной таблицы: строки - "Район", столбцы - "ТипСПросроком", значения - количество строк
    With pivotTable
        .PivotFields("Район").Orientation = xlRowField
        .PivotFields("ТипСПросроком").Orientation = xlColumnField
        .AddDataField .PivotFields("ID нарушения"), "Количество", xlCount
    End With
        With pivotTable
        .GrandTotalName = "Сумма по просрочкам" ' Замените на нужное название для общего итога
    End With
    wsPivot.Range(wsPivot.Cells(pivotStartRow + 1, 1), wsPivot.Cells(pivotStartRow + 1, 1)).Value = "Район" ' Устанавливаем заголовок
    wsPivot.Rows(pivotStartRow).Hidden = True ' Скрываем строку с заголовками сводной таблицы

    ' Убираем столбец "Пусто" (где ТипСПросроком неопределен)
    Dim typePivotField As PivotField
    Set typePivotField = pivotTable.PivotFields("ТипСПросроком")
    For Each item In typePivotField.PivotItems
        If item.Name = "(blank)" Then
            item.Visible = False
        End If
    Next item
    ' Обновляем сводную таблицу
    pivotTable.RefreshTable

    ' Форматирование сводной таблицы
    Dim rng As Range
    Set rng = wsPivot.Range(wsPivot.Cells(pivotStartRow + 1, 1), wsPivot.Cells(pivotStartRow + 1, 1)).CurrentRegion
    With rng
        .Font.Name = "Times New Roman"
        .Font.Size = 14
        .Font.Bold = True
        .Borders.LineStyle = xlContinuous
        .WrapText = True ' Перенос текста
        .HorizontalAlignment = xlCenter ' Выравнивание по центру
        .VerticalAlignment = xlCenter
    End With
    wsPivot.Columns("A").ColumnWidth = 24 ' Установите желаемую ширину столбца
    wsPivot.Rows(pivotStartRow + 1).RowHeight = 53
    wsPivot.Rows(pivotStartRow + 3).RowHeight = 19 ' Установите высоту строки
    wsPivot.Columns("B").ColumnWidth = 39
    wsPivot.Columns("C").ColumnWidth = 34
    wsPivot.Columns("D").ColumnWidth = 33 
    wsPivot.Columns("E").ColumnWidth = 39 

    ' Изменение цвета текста в столбцах, содержащих заданные словосочетания
    Dim col As Integer
    Dim cell As Range
    Dim found As Boolean
    Dim searchStrings As Variant
    searchStrings = Array("В работе с просроком") ' Массив искомых словосочетаний

    For col = 1 To rng.Columns.Count
        found = False
        For Each cell In rng.Columns(col).Cells
            ' Проверяем только строки со 2-й по последнюю (исключая заголовок и итоговые строки)
            If cell.Row > pivotStartRow And cell.Row < rng.Rows.Count + pivotStartRow Then
                If Not IsEmpty(cell.Value) Then
                    For Each searchString In searchStrings
                        If InStr(1, cell.Value, searchString, vbTextCompare) > 0 Then
                            found = True
                            Exit For
                        End If
                    Next searchString
                End If
            End If
            If found Then Exit For
        Next cell

        ' Если найдено, изменить цвет текста в столбце
        If found Then
            ' Изменяем цвет текста только для значений, начиная со 2-й строки
            For Each cell In rng.Columns(col).Cells
                If cell.Row > pivotStartRow + 1 And cell.Row < rng.Rows.Count + pivotStartRow -1 Then
                    cell.Font.Color = RGB(255, 0, 0) ' Красный цвет
                End If
            Next cell
        End If
    Next col
End Sub
"""
    vba_macro_snow = """Sub CreatePivotTableSnow()
        Dim wsData As Worksheet
        Dim wsPivot As Worksheet
        Dim pivotCache As PivotCache
        Dim pivotTable As PivotTable
        Dim lastRow As Long
        Dim lastCol As Long
        Dim pivotStartRow As Long

        ' Укажите лист с данными
        Set wsData = ThisWorkbook.Sheets("СВОД") ' Замените на имя вашего листа с данными

        ' Укажите существующий лист для сводной таблицы
        Set wsPivot = ThisWorkbook.Sheets("Сводная таблица") ' Замените на имя вашего листа с существующей сводной таблицей

        ' Находим последний заполненный ряд и столбец на листе с данными
        lastRow = wsData.Cells(wsData.Rows.Count, "A").End(xlUp).Row
        lastCol = wsData.Cells(1, wsData.Columns.Count).End(xlToLeft).Column

        ' Находим строку, где уже существует сводная таблица, и добавляем новую через 3 строки
        pivotStartRow = wsPivot.Cells(wsPivot.Rows.Count, 1).End(xlUp).Row + 3

        ' Создаем кэш для сводной таблицы
        Set pivotCache = ThisWorkbook.PivotCaches.Create( _
            SourceType:=xlDatabase, _
            SourceData:=wsData.Cells(1, 1).Resize(lastRow, lastCol))

        ' Создаем сводную таблицу
        Set pivotTable = pivotCache.CreatePivotTable( _
            TableDestination:=wsPivot.Cells(pivotStartRow, 1), _
            TableName:="Pivotsnow")

        ' Настройка сводной таблицы: строки - "Район", столбцы - "ТипСнег", значения - количество строк
        With pivotTable
            .PivotFields("Район").Orientation = xlRowField
            .PivotFields("ТипСнег").Orientation = xlColumnField
            .AddDataField .PivotFields("ID нарушения"), "Количество", xlCount
        End With
            With pivotTable
            .GrandTotalName = "Сумма по снегу" ' Замените на нужное название для общего итога
        End With
        wsPivot.Range(wsPivot.Cells(pivotStartRow + 1, 1), wsPivot.Cells(pivotStartRow + 1, 1)).Value = "Район" ' Устанавливаем заголовок
        wsPivot.Rows(pivotStartRow).Hidden = True ' Скрываем строку с заголовками сводной таблицы

        ' Убираем столбец "Пусто" (где ТипСнег неопределен)
        Dim typePivotField As PivotField
        Set typePivotField = pivotTable.PivotFields("ТипСнег")
        For Each item In typePivotField.PivotItems
            If item.Name = "(blank)" Then
                item.Visible = False
            End If
        Next item
        ' Обновляем сводную таблицу
        pivotTable.RefreshTable

        ' Форматирование сводной таблицы
        Dim rng As Range
        Set rng = wsPivot.Range("A39").CurrentRegion
        With rng
            .Font.Name = "Times New Roman"
            .Font.Size = 14
            .Font.Bold = True
            .Borders.LineStyle = xlContinuous
            .WrapText = True ' Перенос текста
            .HorizontalAlignment = xlCenter ' Выравнивание по центру
            .VerticalAlignment = xlCenter
        End With
        wsPivot.Columns("A").ColumnWidth = 24 ' Установите желаемую ширину столбца
        With rng
            .HorizontalAlignment = xlCenter ' Горизонтальное выравнивание по центру
            .VerticalAlignment = xlCenter ' Вертикальное выравнивание по центру
        End With

        ' Настройка высоты строк и ширины столбцов
        wsPivot.Range("40:52").RowHeight = 19
        wsPivot.Columns("B").ColumnWidth = 39
        wsPivot.Columns("C").ColumnWidth = 34
        wsPivot.Columns("D").ColumnWidth = 33
        wsPivot.Columns("E").ColumnWidth = 39

        ' Проверяем наличие столбца с названием, содержащим "снег"
        foundTodayColumn = False
        For Each cell In wsPivot.Range("B37:C39")
            If InStr(1, cell.Value, "Сегодня", vbTextCompare) > 0 Then
                foundTodayColumn = True 
                cell.Font.Color = RGB(255, 0, 0) ' Красный цвет текста заголовка
                Dim dataRange As Range
                Dim lastDataRow As Long
                lastDataRow = wsPivot.Cells(wsPivot.Rows.Count, cell.Column).End(xlUp).Row - 1 ' Уменьшаем на одну строку для исключения итогов
                Set dataRange = wsPivot.Range(cell.Offset(1, 0), wsPivot.Cells(lastDataRow, cell.Column))
                ' Применяем заливку к значениям > 0, исключая итоги
                For Each dataCell In dataRange
                    If IsNumeric(dataCell.Value) And dataCell.Value > 0 Then
                        dataCell.Interior.Color = RGB(247, 134, 126) ' Красная заливка для положительных значений
                    End If
                Next dataCell
            End If
        Next cell
    End Sub
    """

    # Запускаем Excel
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True  # Если нужно, чтобы Excel не отображался, оставьте False

    # Открываем Excel-файл
    workbook = excel.Workbooks.Open(excel_file)

    # Добавляем новый модуль VBA и вставляем макрос
    vb_module = workbook.VBProject.VBComponents.Add(1)  # 1 = стандартный модуль
    vb_module.CodeModule.AddFromString(vba_macro)
    # Выполняем макрос
    excel.Application.Run("CreatePivotTable1")
    print("Pivot1 created")

    vb_module1 = workbook.VBProject.VBComponents.Add(1)  # 1 = стандартный модуль
    vb_module1.CodeModule.AddFromString(vba_macro2)
    excel.Application.Run("CreatePivotTable2")
    print("Pivot2 created")

    if not df[df["Проблема"].isin(["Наличие снега, наледи", "Неочищенная кровля"])].empty:
        vb_module2 = workbook.VBProject.VBComponents.Add(1)  # 1 = стандартный модуль
        vb_module2.CodeModule.AddFromString(vba_macro_snow)
        excel.Application.Run("CreatePivotTableSnow")
        print("CreatePivotTableSnow")

    pdf_file_name = f"Монитор_в_работе_{timenow}_{datetime.now().strftime('%d.%m.%y')}.pdf"
    pdf_path = os.path.join(os.path.dirname(processed_file_path), pdf_file_name)  # Формируем путь к PDF
    wsFirst = workbook.Worksheets(1)  # Ссылка на первый лист

    # Настройки страницы для печати
    wsFirst.PageSetup.FitToPagesWide = 1  # Устанавливаем количество страниц по ширине
    wsFirst.PageSetup.FitToPagesTall = 1  # Устанавливаем количество страниц по высоте на 1
    wsFirst.PageSetup.Zoom = False  # Отключаем масштабирование

    # Обновляем отступы страницы для уменьшения размера PDF
    wsFirst.PageSetup.LeftMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.RightMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.TopMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.BottomMargin = excel.Application.CentimetersToPoints(0.5)
    workbook.Save()
    try:
        # Убираем ошибку, если файл уже существует
        if os.path.exists(pdf_path):
            print(f"Файл {pdf_path} существует. Удаление...")
            os.remove(pdf_path)  # Удаляем файл, если он существует
            print("Файл успешно удален.")

        print(f"Сохранение файла в {pdf_path}...")
        wsFirst.ExportAsFixedFormat(0, pdf_path)  # 0 - это xlTypePDF
        print(f"PDF успешно создан: {pdf_path}")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
    sheet = workbook.Worksheets(2)
    sheet.Cells.EntireColumn.AutoFit()
    # Сохраняем и закрываем файл
    workbook.Save()
    workbook.Close()

    # Закрываем Excel
    excel.Quit()
    return processed_file_path, pdf_path


# ЛК Префекта -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
async def call_lk_prefekta(update: Update,chat_id, context: CallbackContext, district: str) -> None:
    success = await parcing_data_lk_prefekta(context, chat_id)  # Передаем контекст и ID чата
    if not success:
        return  # Если произошла ошибка, выходим из обработчика
    files = os.listdir(directory)
    files.sort(key=lambda x: os.path.getmtime(os.path.join(directory, x)))
    latest_downloaded_file = files[-1]
    filepath = os.path.join(directory, latest_downloaded_file)
    processed_file_path = process_lk_prefekta_file(directory, district, filepath)
    if not processed_file_path:
        error_message = f"❌ Заявок ЛК Префекта по данному району нет!"
        print(error_message)  # Выводим ошибку в консоль
        await context.bot.send_message(chat_id=chat_id, text=error_message)  # Отправляем сообщение в Telegram
        return #выходим из обработки
    else:
        with open(processed_file_path, 'rb') as f:
            await update.callback_query.message.reply_document(InputFile(f))
async def parcing_data_lk_prefekta(context, chat_id):
    chrome_install = ChromeDriverManager().install()
    folder = os.path.dirname(chrome_install)
    chromedriver_path = os.path.join(folder, "chromedriver.exe")
    driver = webdriver.Chrome(service=ChromeService(chromedriver_path))
    try:
        # Откройте страницу логина
        driver.get('https://gorod.mos.ru/api/service/auth/auth')

        # Найдите поля для ввода логина и пароля и заполните их
        username = driver.find_element(By.XPATH, '//input[@placeholder="Логин *"]')
        password = driver.find_element(By.XPATH, '//input[@placeholder="Пароль*"]')
        username.send_keys(login_NG)
        password.send_keys(password_NG)

        # Найдите и нажмите кнопку логина
        login_button = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div[2]/form[1]/button')
        login_button.click()
        # Подождите, пока страница загрузится
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,
                                                                        '//div[@class="dashboard__block-link"]//div[@class="button-big link"]//div[@class="dashboard-container__links-title" and contains(text(), "Аналитика")]')))
        # переход в ответы в работе
        driver.get('https://gorod.mos.ru/admin/ker/olap/report/155')
        time.sleep(10)
        # # прыжок в меню
        # button = driver.find_element(By.XPATH,
        #                              "/html/body/div[3]/div/div[2]/div/div/div/div/form/header/div[1]/button[1]/span[2]/i")
        # button.click()
        # time.sleep(4)
        # # выбор фильтра
        # WebDriverWait(driver, 20).until(EC.presence_of_element_located(
        #     (By.XPATH, '/html/body/div[3]/div/div[2]/div/div/div/div/form/div[1]/aside/div/div[2]/div/div[1]/div/a')))
        # button = driver.find_element(By.XPATH,
        #                              "/html/body/div[3]/div/div[2]/div/div/div/div/form/div[1]/aside/div/div[2]/div/div[1]/div/a")
        # button.click()

        # экспорт
        WebDriverWait(driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span')))
        button = driver.find_element(By.XPATH,
                                     '/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span')
        button.click()
        time.sleep(1)

        # # ок- выгркзка с экселя
        # button = driver.find_element(By.XPATH, '/html/body/div[4]/div/div[2]/div/div[3]/button[2]/span[2]/span')
        # button.click()
        # time.sleep(1)

        # one more time click to export
        button = driver.find_element(By.XPATH, "//button[contains(@class, 'bg-primary')]//span[text()='Экспорт']")
        button.click()
        time.sleep(1)

        # переход в загрузки
        driver.get('https://gorod.mos.ru/admin/ker/olap/downloads')
        # Подождите, пока страница загрузится)
        WebDriverWait(driver, 1500).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')))
        # скачивание файла
        button = driver.find_element(By.XPATH,
                                     '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')
        button.click()
        time.sleep(20)
        return True
    except Exception as e:
        error_message = f"❌Произошла ошибка при выгрузке ЛК префекта. Пожалуйста, попробуйте еще раз."
        print(error_message)  # Выводим ошибку в консоль
        await context.bot.send_message(chat_id=chat_id, text=error_message)  # Отправляем сообщение в Telegram
        await context.bot.send_message(chat_id=chat_id, text='Выберите команду:',
                                       reply_markup=InlineKeyboardMarkup([
                                           [InlineKeyboardButton("🏢 ЛК префекта(НГ)", callback_data='lk_prefekt')],
                                           [InlineKeyboardButton("📊 Монитор в Работе(ММ)", callback_data='mm_monitor')],
                                           [InlineKeyboardButton("📈 Ответы в работе (НГ)", callback_data='ng_answers')],
                                           [InlineKeyboardButton("📋 СВОД МЖИ(НГ)", callback_data='mji_summary')],
                                           [InlineKeyboardButton("📅 Статистика МЖИ", callback_data='mji_stat')],
                                           [InlineKeyboardButton("🌐 Монитор (НГ)", callback_data='city_monitor')],
                                           [InlineKeyboardButton("♻️ КП/БП (ММ)", callback_data='mm_kp_bp')],
                                           [InlineKeyboardButton("🚨 Ежедневные просрочки (ММ)",
                                                                 callback_data='today_mm')],
                                           [InlineKeyboardButton("🔄 Последнее обновление",
                                                                 callback_data='last_update')],
                                           [InlineKeyboardButton("❓ Объяснение команд", callback_data='explain')],
                                       ]))
        return False
    finally:
        driver.quit()
def process_lk_prefekta_file(directory: str, selected_district: str, filepath: str) -> str:
    df = pd.read_excel(filepath)

    responsible_mapping = {
        'ГБУ «Автомобильные дороги ЮВАО»': 'АВД ЮВАО',
        'ГБУ Жилищник Выхино района Выхино-Жулебино города Москвы': 'Выхино-Жулебино',
        'Управа Выхино-Жулебино': 'Выхино-Жулебино',
        'ГБУ Жилищник Нижегородского района города Москвы': 'Нижегородский',
        'Управа Нижегородский': 'Нижегородский',
        'ГБУ Жилищник района Капотня города Москвы': 'Капотня',
        'Управа Капотня': 'Капотня',
        'ГБУ Жилищник района Кузьминки города Москвы': 'Кузьминки',
        'Управа Кузьминки': 'Кузьминки',
        'ГБУ Жилищник района Лефортово города Москвы': 'Лефортово',
        'Управа Лефортово': 'Лефортово',
        'ГБУ Жилищник района Люблино города Москвы': 'Люблино',
        'Управа Люблино': 'Люблино',
        'ГБУ Жилищник района Марьино города Москвы': 'Марьино',
        'Управа Марьино': 'Марьино',
        'ГБУ Жилищник района Некрасовка города Москвы': 'Некрасовка',
        'Управа Некрасовка': 'Некрасовка',
        'ГБУ Жилищник района Печатники города Москвы': 'Печатники',
        'Управа Печатники': 'Печатники',
        'ГБУ Жилищник района Текстильщики города Москвы': 'Текстильщики',
        'Управа Текстильщики': 'Текстильщики',
        'ГБУ Жилищник Рязанского района города Москвы': 'Рязанский',
        'Управа Рязанский': 'Рязанский',
        'ГБУ Жилищник Южнопортового района города Москвы': 'Южнопортовый',
        'Управа Южнопортовый': 'Южнопортовый'
    }

    # Функция для обновления значений в столбце 'Район'
    def update_region(row):
        if row['Ответственный ОИВ первого уровня'] == 'Префектура Юго-Восточного округа':
            return row['Район']  # Ничего не меняем
        else:
            return responsible_mapping.get(row['Ответственный ОИВ первого уровня'], row['Район'])

    # Применение функции к каждому ряду
    df['Район'] = df.apply(update_region, axis=1)

    df_filtered = df[df['Ответственный за подготовку ответа'] == 'Префектура Юго-Восточного округа']

    columns_to_keep = [
        "Номер заявки",
        "Регламентный срок у сообщения (Портал)",
        "Дата публикации сообщения",
        "Район",
        "Проблемная тема",
        "Адрес",
        "Категория объекта",
        "Категория/действие последнего ответа",
        "Ответственный за подготовку ответа",
        "Ответственный ОИВ первого уровня",
        "Статус подготовки ответа на сообщение"
    ]
    df_filtered = df_filtered[columns_to_keep]
    if selected_district != "Все районы":
        df_filtered = df_filtered[df_filtered['Район'] == selected_district]

    # Удаляем полностью пустые строки и проверяем количество оставшихся строк
    df_filtered = df_filtered.dropna(how='all')
    if df_filtered.empty:
        return False

    now = pd.Timestamp.now()
    processed_file_path = os.path.join(directory,
                                       f"{selected_district}_ЛК_Префекта_{datetime.now().strftime('%d.%m')}_на_{now.strftime('%H-%M')}.xlsx")
    print(f"Saving processed file to: {processed_file_path}")
    df_filtered.to_excel(processed_file_path, index=False)
    excel_file = processed_file_path
    vba_macro = """  
            Sub CreatePivotTable()  
                Dim wsData As Worksheet  
                Dim wsPivot As Worksheet  
                Dim pivotCache As PivotCache  
                Dim pivotTable As PivotTable  
                Dim lastRow As Long  
                Dim lastCol As Long  

                ' Укажите лист с данными  
                Set wsData = ThisWorkbook.Sheets("Sheet1") ' Замените на имя вашего листа с данными  

                With wsData.Columns("B")
                    .NumberFormat = "DD.MM.YYYY"
                End With

                ' Создаем новый лист для сводной таблицы  
                On Error Resume Next  
                Application.DisplayAlerts = False  
                ThisWorkbook.Sheets("Сводная таблица").Delete ' Удаляем лист, если уже существует  
                Application.DisplayAlerts = True  
                On Error GoTo 0  
                Set wsPivot = ThisWorkbook.Sheets.Add  
                wsPivot.Name = "Сводная таблица"  

                ' Находим последний заполненный ряд и столбец на листе с данными  
                lastRow = wsData.Cells(wsData.Rows.Count, "A").End(xlUp).Row  
                lastCol = wsData.Cells(1, wsData.Columns.Count).End(xlToLeft).Column  

                ' Создаем кэш для сводной таблицы  
                Set pivotCache = ThisWorkbook.PivotCaches.Create( _  
                    SourceType:=xlDatabase, _  
                    SourceData:=wsData.Cells(1, 1).Resize(lastRow, lastCol))  

                ' Создаем сводную таблицу  
                Set pivotTable = pivotCache.CreatePivotTable( _  
                    TableDestination:=wsPivot.Cells(3, 1), _  
                    TableName:="MyPivotTable")  

                With pivotTable  
                    .PivotFields("Район").Orientation = xlRowField  
                    .PivotFields("Регламентный срок у сообщения (Портал)").Orientation = xlColumnField  
                    .AddDataField .PivotFields("Номер заявки"), "Количество", xlCount  
                End With  

                wsPivot.Range("A4").Value = "Район" 
                ' Скрываем первую строку  
                wsPivot.Rows(3).Hidden = True  

                ' Обновляем сводную таблицу  
                pivotTable.RefreshTable  

                ' Форматирование сводной таблицы  
                Dim rng As Range  
                Set rng = wsPivot.Range("A4").CurrentRegion  
                With rng  
                    .Font.Name = "Times New Roman"  
                    .Font.Size = 14  
                    .Font.Bold = True  
                    .Borders.LineStyle = xlContinuous  
                    .WrapText = True ' Перенос текста  
                    .HorizontalAlignment = xlCenter ' Выравнивание по центру  
                End With  
                wsPivot.Columns("A").ColumnWidth = 24 ' Установите желаемую ширину столбца  
                wsPivot.Rows(6).RowHeight = 19 ' Установите высоту 6-й строки

            End Sub
            """
    # Запускаем Excel
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True  # Если нужно, чтобы Excel не отображался, оставьте False

    # Открываем Excel-файл
    workbook = excel.Workbooks.Open(excel_file)

    # Добавляем новый модуль VBA и вставляем макрос
    vb_module = workbook.VBProject.VBComponents.Add(1)  # 1 = стандартный модуль
    vb_module.CodeModule.AddFromString(vba_macro)

    # Выполняем макрос
    excel.Application.Run("CreatePivotTable")
    print("Pivot created")
    pdf_file_name = f"{selected_district}_ЛК_Префекта_{datetime.now().strftime('%d.%m')}_на_{now.strftime('%H-%M')}.xlsx"
    pdf_path = os.path.join(os.path.dirname(processed_file_path), pdf_file_name)  # Формируем путь к PDF
    wsFirst = workbook.Worksheets(1)  # Ссылка на первый лист

    # Настройки страницы для печати
    wsFirst.PageSetup.FitToPagesWide = 1  # Устанавливаем количество страниц по ширине
    wsFirst.PageSetup.FitToPagesTall = 1  # Устанавливаем количество страниц по высоте на 1
    wsFirst.PageSetup.Zoom = False  # Отключаем масштабирование

    # Обновляем отступы страницы для уменьшения размера PDF
    wsFirst.PageSetup.LeftMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.RightMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.TopMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.BottomMargin = excel.Application.CentimetersToPoints(0.5)
    workbook.Save()
    try:
        # Убираем ошибку, если файл уже существует
        if os.path.exists(pdf_path):
            print(f"Файл {pdf_path} существует. Удаление...")
            os.remove(pdf_path)  # Удаляем файл, если он существует
            print("Файл успешно удален.")

        print(f"Сохранение файла в {pdf_path}...")
        wsFirst.ExportAsFixedFormat(0, pdf_path)  # 0 - это xlTypePDF
        print(f"PDF успешно создан: {pdf_path}")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
    sheet = workbook.Worksheets(2)
    sheet.Cells.EntireColumn.AutoFit()
    # Сохраняем и закрываем файл
    workbook.Save()
    workbook.Close()

    # Закрываем Excel
    excel.Quit()
    return processed_file_path